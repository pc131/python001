#test script that extraxts test data from excel files (prepared by Katherine)
#and creates python dictionary in the style of: {'3019601851S1X':('Elster',''88025763'),'3019601916S11':('Elster',''93A015953'),SPID:(METER_MNF, METER_SER)}
#which dictionary is ater used in generate_test_case_ALL.py script in SPIDS_METERS variable/constant
#SPIDS_METERS = {'3019601851S1X':('Elster',''88025763'),'3019601916S11':('Elster',''93A015953'),SPID:(METER_MNF, METER_SER),.........}

import random
import openpyxl as xl

working_dir = "C:\\Users\\tomasz.skoczylas\\Downloads\\11\\"
spids_meters_filename = working_dir + "Roy_Test Data Release6_assurance.xlsx"
wb1 = xl.load_workbook(spids_meters_filename)
# ws11 = wb1.worksheets[0]
# ws12 = wb1.worksheets[1]


standalone_spids = wb1.worksheets[2]

def pick_spid_meter_xlsx():
    # how many different SPIDS and METERS to pick from the Excel
    dict_spids = "{"
    for a in range(2,42): #choose SPIDs from range of rows
        spid = standalone_spids.cell(row=a, column=1).value
        meter_mnf = standalone_spids.cell(row=a, column=3).value
        meter_ser = standalone_spids.cell(row=a, column=4).value
        dict_spids += "'" + str(spid) + "':('" + str(meter_mnf) + "','" + str(meter_ser) + "'),"
    dict_spids = dict_spids[:-1] # remove last comma  
    dict_spids += "}"
    print(dict_spids)
    #pyperclip.copy(dict_spids)
pick_spid_meter_xlsx()