import openpyxl as xl
from openpyxl.styles import PatternFill
import os
from datetime import datetime
import random
import string
import names
import sys
from random_address import real_random_address
from faker import Faker

working_dir = "C:\\Users\\tomasz.skoczylas\\Downloads\\11\\"
filename = working_dir + "Bilaterals 1.3.3.1 master.xlsx"
wb1 = xl.load_workbook(filename)
ws11 = wb1.worksheets[0]
ws12 = wb1.worksheets[1]

RETAILER = 'MOSLTEST-R'
WHOLESALER = 'MOSLTEST-W'
T216_URL = 'https://moservicesdev.mosl.co.uk/test/attachments/87ffc85e-ebd5-461c-99d6-2ac3eef43f7c'

# TEST_CASE_SEQUENCE = ['T501.R'] # SUBMITTED
# TEST_CASE_SEQUENCE = ['T501.R', 'T201.W'] #ACCEPTED
TEST_CASE_SEQUENCE = ['T501.R', 'T201.W', 'T213.W'] #DEFERRED !!! CHECK THE DATES of T213W_data_items
# TEST_CASE_SEQUENCE = ['T501.R', 'T211.R'] #CANCELLED
# TEST_CASE_SEQUENCE = ['T501.R', 'T202.W'] #REJECTED
# TEST_CASE_SEQUENCE = ['T501.R', 'T202.W', 'T210.R'] #RESUBMITTED 
# TEST_CASE_SEQUENCE = ['T501.R', 'T201.W', 'T203.W'] #INFOREQST
# TEST_CASE_SEQUENCE = ['T501.R', 'T201.W', 'T203.W', 'T204.R'] # INFOPROVD
# TEST_CASE_SEQUENCE = ['T501.R', 'T201.W', 'T222.W'] # COMPLETED
# TEST_CASE_SEQUENCE = ['T501.R', 'T201.W', 'T203.W', 'T204.R', 'T222.W', 'T208.R'] #CLOSED
# TEST_CASE_SEQUENCE = ['T501.R', 'T501.R', 'T201.W', 'T501.R', 'T211.R', 'T501.R', 'T202.W', 'T501.R', 'T202.W', 'T210.R', 'T501.R', 'T201.W', 'T203.W', 'T501.R', 'T201.W', 'T222.W']

TEST_CASE_LENGTH = len(TEST_CASE_SEQUENCE)

# SPIDS and METERS can be selected straight from Excel Test Data
# spids_meters_filename = working_dir + "TEST_DATA.xlsx"
# wb2 = xl.load_workbook(spids_meters_filename)
# standalone_spids = wb2.worksheets[1]

POSTCODES = ['B1 1HQ', 'BN88 1AH', 'BS98 1TL', 'BX1 1LT', 'BX2 1LB', 'BX3 2BB', 'BX4 7SB', 'BX5 5AT', 'CF10 1BH', 'CF99 1NA', 'CO4 3SQ', 'CV4 8UW', 'CV35 0DB', 'E14 5EY', 'DA1 1RT', 'DE99 3GG', 'DE55 4SW', 'DH98 1BT', 'DH99 1NS', 'E14 5HQ', 'E14 5JP', 'E16 1XL', 'E20 2AQ', 'E20 2BB', 'E20 2ST', 'E20 3BS', 'E20 3EL', 'E20 3ET', 'E20 3HB', 'E20 3HY', 'E98 1SN', 'E98 1ST', 'E98 1TT', 'EC2N 2DB', 'EC4Y 0HQ', 'EH12 1HQ', 'EH99 1SP', 'G58 1SB', 'GIR 0AA', 'IV21 2LR', 'L30 4GB', 'LS98 1FD', 'M50 2BH', 'M50 2QH', 'N1 9G', 'N81 1ER', 'NE1 4ST', 'NG80 1EH', 'NG80 1LH', 'NG80 1RH', 'NG80 1TH', 'PH1 5RB', 'PH1 2SJ', 'S2 4SU	', 'S6 1SW', 'S14 7UP', 'SE1 0NE', 'SE1 8UJ', 'SM6 0HB', 'SN38 1NW', 'SR5 1SU', 'SW1A 0AA', 'SW1A 0PW', 'SW1A 1AA', 'SW1A 2AA', 'SW1A 2AB', 'SW1H 0TL', 'SW1P 3EU', 'SW1W 0DT', 'SW11 7US', 'SW19 5AE', 'TW8 9GS', 'W1A 1AA', 'W1D 4FA', 'W1N 4DJ', 'W1T 1FB']
SPIDS_METERS = {'3019053153W15':('ARAD','9125420'),'3019053676W14':('ELSTER','8121165'),'3019053781W12':('NEPTUNE_MEASUREMENT','273298'),'3019053803W15':('KENT','82122903'),'3019053811W12':('KENT','82095722'),'301905382XW1X':('ARAD','9317245'),'3019054370W16':('ARAD','8103835'),'3019054419W13':('ARAD','8175760'),'3019054672W19':('SCHLUMBERGER','98AQ363729'),'3019054869W13':('ARAD','15A10272'),'3019055199W13':('ARAD','8005930'),'3019055210W16':('KENT','3M320765'),'3019055431W12':('KENT','4A030096'),'3019055482W19':('SCHLUMBERGER','98AQ328599'),'3019055490W16':('ARAD','9034613'),'3019055733W15':('ARAD','8092051'),'3019055768W17':('KENT','576456'),'3019055784W11':('NEPTUNE_MEASUREMENT','PB060105'),'3019055814W11':('KENT','82056507'),'3019055865W18':('ARAD','9131620'),'3019055970W16':('ARAD','9113629'),'3019056020W16':('SOCAM','4444369'),'3019056217W10':('SCHLUMBERGER','97AQ219008'),'3019056322W19':('KENT','4A059262'),'3019056403W15':('KENT','429332'),'3019056470W16':('KENT','82044656'),'3019056713W15':('KENT','5M049553'),'3019056772W19':('KENT','3A246197'),'3019056829W13':('AQUADIS','96PA186385'),'3019057000W16':('SW_METER','8M044268'),'3019057035W18':('KENT','80229964'),'3019057655W18':('ARAD','11A10114'),'3019058147W10':('KENT','4M093691'),'3019058201W12':('ARAD','8405271'),'3019058309W13':('ARAD','8159546'),'3019058546W14':('ARAD','16A10038'),'3019058988W17':('KENT','4T033142'),'301905950XW1X':('KENT','93M178899'),'3019059623W15':('KENT','405042'),'301906001XW1X':('KENT','91P701223'),'3019060729W13':('SW_METER','9M129944'),'3019060796W14':('KENT','93A028514'),'3019060834W11':('KENT','3A250702'),'3019060931W12':('SW_METER','9M268716'),'3019061105W18':('NEPTUNE_MEASUREMENT','5M312526'),'3019061113W15':('ARAD','8145048'),'3019061164W11':('KENT','4A219424'),'3019061350W16':('ELSTER','12M50046'),'3019061636W14':('ATLANTIC_PLASTIC','875165'),'3019061660W16':('ARAD','8277392'),'3019062039W13':('SW_METER','8A095703'),'3019062276W14':('SW_METER','8M000312'),'3019062802W19':('KENT','93M151324'),'3019062845W18':('SCHLUMBERGER','98AQ368809'),'3019063213W15':('ELSTER','8088420'),'301906323XW1X':('ELSTER','10A10289'),'3019063248W17':('ARAD','9048511'),'3019063329W13':('ARAD','8969925'),'3019063809W13':('SW_METER','8M026193'),'3019063957W10':('KENT','85203324'),'3019063981W12':('KENT','4A016617'),'301906399XW1X':('KENT','10A011901'),'3019064031W12':('SCHLUMBERGER','4M043165'),'3019064295W18':('KENT','4A141106'),'301906435XW1X':('KENT','89163811'),'3019064457W10':('KENT','89074837'),'3019064473W15':('ARAD','9074723'),'3019065003W15':('SW_METER','8M482077'),'3019065402W19':('ELSTER','7A199413'),'3019065437W10':('KENT','4A219416'),'3019065593W15':('ELSTER','8058283'),'3019065631W12':('KENT','82041988'),'3019065682W19':('KENT','82333230'),'3019065704W11':('NEPTUNE_MEASUREMENT','91P020049'),'3019065720W16':('KENT','429371'),'3019065836W14':('ARAD','8217903'),'3019065844W11':('ARAD','8440908'),'301906614XW1X':('KENT','90103896'),'3019066360W16':('KENT','4A016848'),'301906645XW1X':('ARAD','8288023'),'3019066565W18':('KENT','90P184000'),'3019066611W12':('ELSTER','12A10167'),'3019066751W12':('SCHLUMBERGER','97AQ286279'),'3019066778W17':('KENT','92A627036'),'3019066816W14':('KENT','3M246967'),'3019066867W10':('SCHLUMBERGER','96AQ186751'),'3019066964W11':('KENT','94A013957'),'3019067138W17':('AQUADIS','97AQ295713'),'3019067197W10':('ARAD','8119732'),'3019067413W15':('KENT','3M032653'),'3019067456W14':('AQUADIS','98PA082002'),'3019067472W19':('KENT','83277294'),'3019067510W16':('KENT','2A112403'),'3019067537W10':('KENT','73261026'),'3019067715W18':('KENT','396287'),'3019067812W19':('KENT','72158247'),'3019067820W16':('KENT','4A059224'),'3019067847W10':('ELSTER','8073665'),'3019067944W11':('ARAD','9118870'),'3019067979W13':('ELSTER','8025637'),'3019068118W17':('ARAD','8307041'),'3019068258W17':('ELSTER','8060566'),'3019068541W12':('ARAD','9072452'),'301906855XW1X':('KENT','92P003826'),'3019068711W12':('ELSTER','9T015564'),'3019068746W14':('SW_METER','7M456837'),'3019068940W16':('KENT','91M121269'),'3019068967W10':('KENT','4M261717'),'3019069106W14':('KENT','4M117004'),'3019069254W11':('ARAD','9066103'),'3019069262W19':('SCHLUMBERGER','97AQ246008'),'3019069491W12':('KENT','99AQ424349'),'3019069629W13':('SCHLUMBERGER','4M098835'),'3019069912W19':('ELSTER','8024675'),'3019069971W12':('KENT','9118152'),'3019070090W16':('KENT','91030736'),'3019070406W14':('KENT','1T008043'),'3019070414W11':('SW_METER','9M369341'),'3019070430W16':('SW_METER','06A106428'),'3019070465W18':('ELSTER','8599442'),'3019070627W10':('KENT','92M073908'),'3019070732W19':('KENT','95M186415'),'3019070759W13':('ARAD','8136617'),'3019071208W17':('ABB','3A082107'),'3019071267W10':('ARAD','8998326'),'3019071283W15':('KENT','4M256224'),'3019071372W19':('KENT','85224222'),'3019071380W16':('KENT','72204671'),'3019071437W10':('ARAD','9135615'),'3019071488W17':('KENT','1128505'),'301907150XW1X':('SCHLUMBERGER','96AQ105006'),'3019071747W10':('KENT','4A023469'),'3019071763W15':('ARAD','9510533'),'3019071860W16':('KENT','94M190521'),'3019071941W12':('KENT','4T001182'),'3019072271W12':('ARAD','8261279'),'3019072298W17':('KENT','97PA083061'),'3019072530W16':('KENT','93A008413'),'3019072735W18':('KENT','83276871'),'3019072832W19':('ELSTER','8512334'),'3019073235W18':('KENT','AG530317'),'3019073359W13':('KENT','4M117431'),'3019073715W18':('ARAD','8077151'),'3019073987W10':('SW_METER','8M064134'),'3019074150W16':('ARAD','8005921'),'3019074193W15':('KENT','4T007900'),'3019074215W18':('KENT','4A065728'),'3019074339W13':('ARAD','18AI0084'),'301907438XW1X':('KENT','91M126574'),'301907441XW1X':('ELSTER','8068156'),'3019074568W17':('ARAD','9120181'),'3019074924W11':('KENT','3A082104'),'3019075203W15':('ARAD','13225774'),'3019075378W17':('KENT','88126018'),'3019075386W14':('KENT','4T018783'),'3019075491W12':('SW_METER','7W222353'),'3019075599W13':('KENT','222009'),'3019075629W13':('ARAD','20MS000326'),'3019075637W10':('ELSTER','143037460'),'3019075645W18':('KENT','4A094408'),'3019075750W16':('ELSTER','6T013226'),'3019075785W18':('KENT','552520'),'3019075815W18':('ELSTER','8046612'),'301907603XW1X':('SCHLUMBERGER','D1-98AQ334240'),'3019076315W18':('KENT','93A606364'),'3019076390W16':('NEPTUNE_MEASUREMENT','339845'),'3019076498W17':('ABB','92P024308'),'3019076633W15':('KENT','94M095976'),'3019076692W19':('ELSTER','9M150213'),'3019076757W10':('KENT','4T008990'),'3019076773W15':('KENT','367030'),'301907682XW1X':('KENT','4A065723'),'3019076870W16':('ELSTER','6A185907'),'3019077079W13':('KENT','5A084531'),'3019077141W12':('ELSTER','8986277'),'3019077281W12':('AQUADIS','99AQ489220'),'3019077303W15':('ARAD','9156898'),'3019077389W13':('AMR','10MS0020'),'3019077397W10':('KENT','94A023990'),'3019077702W19':('KENT','93A024152'),'3019078342W19':('KENT','93A723373'),'3019078423W15':('KENT','4T006564'),'3019078431W12':('KENT','2T002933'),'3019078482W19':('KENT','92A626982'),'3019078512W19':('KENT','86081622'),'3019078539W13':('KENT','86081621'),'3019078563W15':('KENT','4A055331'),'3019078598W17':('SCHLUMBERGER','98AQ374259'),'3019078695W18':('ARAD','8351894'),'3019078806W14':('KENT','4M135908'),'3019078954W11':('SW_METER','6A200109'),'3019078989W13':('KENT','4M290996'),'3019079071W12':('ARAD','19A10017'),'3019079446W14':('KENT','AL022622'),'3019079462W19':('KENT','82095907'),'3019079470W16':('KENT','4A023506'),'3019079691W12':('ARAD','8250373'),'3019079705W18':('ELSTER','91S000609'),'3019079713W15':('KENT','6A125902')}

# def pick_spid_meter_xlsx():
#     spids = []
#     meters = []
#     # how many different SPIDS and METERS to pick from the Excel
#     for a in range(20):
#         # define maximum numbers of ROWS in the Sheet
#         row_number = random.randint(2,200)
#         spid = standalone_spids.cell(row=row_number, column=1).value
#         meter_mnf = standalone_spids.cell(row=row_number, column=4).value
#         meter_ser = standalone_spids.cell(row=row_number, column=5).value
#         #appends SPIDS to 1st list
#         spids.append(spid)
#         #append list of meter_mnf and meter_ser as second list
#         meters.append([meter_mnf, meter_ser])
#         #combine these 2 lists as dictionaty, where element from 1st list is key, and second list meter_mnf_, meter_ser is value
#         spids_meters = dict(zip(spids, meters))
#     spid_meter = random.choice(list(spids_meters.items()))
#     spid = spid_meter[0]
#     meter_mnf = spid_meter[1][0]
#     meter_ser = spid_meter[1][1]
#     return spid, meter_mnf, meter_ser

#generate fake data
fake = Faker()

#pick random SPID, METER_MNF_ METER_SERIAL
def pick_spid():
    spids = random.choice(list(SPIDS_METERS.items()))
    spid = spids[0]
    return spid

def random_email():
    return fake.company_email()

def random_string():
    return ''.join(random.choice(string.ascii_letters) for _ in range(15))

def random_name():
    return names.get_full_name()

def random_phone():
    return random.randint(4400000000, 4499999999)

D8036 = ['ERROR', 'DUPLICATE', 'SWITCHED', 'REJECTION', 'UNABLEASST', 'DISAGREEPLAN'] # T211.R T211.W Cancellation Reason Code
D8226 = ['NOCONTACT', 'UNCOOPCUST', 'INACCONTACT', 'MOREDETAILS'] # T203.W T217W Additional Information Request Code
D8228 = ['WHOL', 'NONWHOL'] # T206.W Site Visit Failure Code
D8229 = ['CUSTOMER', 'RETAILER', 'THIRDPARTY', 'CONSENTS', 'REGULAT', 'WEATHER', 'FORCEMAJ', 'INFOREQD'] # T213.W Request Deferral Code
D8230 = ['INACCURATE', 'DUPLICATE', 'WRONGPRO', 'POLICY', 'HOUSEHOLD', 'NOTWHOL'] # T202.W Reject Reason Code
D8231 = ['DISPREJECT', 'DISPCMOS'] # T210.R Resubmit Reason Code
D8236 = ['EMAIL', 'TEL', 'BOTH'] # T321.R T321.W Customer Preferred Method of Contact
D2005 = ['SEMDV', 'NA'] # T321.R T321.W 351.R 351.W Customer Classification – Sensitive Customer
D8237 = ['AM', 'PM', 'BOTH'] # T321.R T321.W T351.R T351.W
D8242 = ['PDF', 'JPG', 'PNG'] # T215.R T215.W

D8356 = ['FIRST', 'FURTHER', 'CCWLEVEL', 'ADR', 'OTHER'] #F5 T501.R Complaint Level
D8358 = ['ADMINISTRATION', 'METERINGASSET', 'BILLING', 'WATER', 'SEWERAGE', 'OTHER'] #F5 T501.R Complaint Category
D8360 = ['GSSFAILURE', 'OTHER', 'NONE'] #F5 T501.R Compensation Claimed
 
def generate_test_case_F5R(loop_times):
    new_filename = ''
    for a in range(loop_times):
        # assign random SPID, METER_MNF, METER_SERIAL to variables - use EXCEL or SPIDS_METERS static dictionary
        SPID = pick_spid()
        #SPID, METER_MNF, METER_SER = pick_spid_meter_xlsx()
        CUST_EMAIL = random_email()
        RET_EMAIL = random_email()
        RANDOM_STRING = random_string()
        CUST_RANDOM_NAME = random_name()
        CUST_RANDOM_PHONE = random_phone()
        RET_RANDOM_NAME = random_name()
        RET_RANDOM_PHONE = random_phone()
        CUST_RANDOM_NAME2 = random_name()
        CUST_RANDOM_PHONE2 = random_phone()
        RET_RANDOM_NAME2 = random_name()
        RET_RANDOM_PHONE2 = random_phone()

        T501R_data_items = [# basic data
                            SPID, 'RET_' + RANDOM_STRING, '[today-' + str(random.randint(0, 7))  +']', fake.paragraph(nb_sentences=1), random.choice(D8356), RANDOM_STRING, random.choice(D8358), RANDOM_STRING,random.choice(D8360), RANDOM_STRING, '', '1', fake.paragraph(nb_sentences=1), '[today]', '1',       # [today - 0] = [today]!!!
                            # customer and retailer data
                            CUST_RANDOM_NAME, CUST_RANDOM_PHONE, '105', CUST_RANDOM_NAME2, CUST_RANDOM_PHONE2, '122', CUST_EMAIL, '1', 'EMAIL', random.choice(D8237),
                            fake.paragraph(nb_sentences=1), random.choice(D2005), fake.paragraph(nb_sentences=1), RET_RANDOM_NAME, RET_RANDOM_PHONE, '210', RET_RANDOM_NAME2, RET_RANDOM_PHONE2, '224', RET_EMAIL,
                            ]
        T201W_data_items = ['[orid]', 'ACCEPTED']
        T202W_data_items = ['[orid]', 'WSL-123456', random.choice(D8230), 'REJECTED']
        T203W_data_items = ['[orid]', random.choice(D8226), 'INFOREQST']
        T204R_data_items = ['[orid]', 'INFOPROVD']
        T207R_data_items = ['[orid]', 'RETAILER_COMMENT']
        T207W_data_items = ['[orid]', 'WHOLSALER_COMMENT']
        T208R_data_items = ['[orid]', 'CLOSED']
        T210R_data_items = ['[orid]', random.choice(D8231), 'RESUBMITTED']
        T211R_data_items = ['[orid]', random.choice(D8036), 'RTL CANCELLED']
        T213W_data_items = ['[orid]', random.choice(D8229), '[today]', '[today+15]', 'START_DEFERRAL']
        T214W_data_items = ['[orid]', '[today]', 'END_DEFERRAL'] # can think of function to peek working day
        T215R_data_items = ['[orid]', '', 'img1png', 'PNG', '4oCwUE5HChoKICAgCklIRFIgICADICAgAwgCICAgxa5KIsSNICAgCXBIWXMgIA7DhCAgDsOEAeKAoisOGyAgICdJREFUCOKEomPDlG7Dn8O2y5nLmX8BNnbCpn/LmcWjMTMzH8OrxI9nYmXLmXd5w7cmxLrFmBAgxZDFnwrFpH4uJsKsICAgIElFTkTCrkJg4oCa']
        T215W_data_items = ['[orid]', '', 'img1png', 'PNG', '4oCwUE5HChoKICAgCklIRFIgICADICAgAwgCICAgxa5KIsSNICAgCXBIWXMgIA7DhCAgDsOEAeKAoisOGyAgICdJREFUCOKEomPDlG7Dn8O2y5nLmX8BNnbCpn/LmcWjMTMzH8OrxI9nYmXLmXd5w7cmxLrFmBAgxZDFnwrFpH4uJsKsICAgIElFTkTCrkJg4oCa']
        T216R_data_items = ['[orid]', T216_URL]
        T216W_data_items = ['[orid]', T216_URL]

        T222W_data_items = [# basic data
                            '[orid]', 'FOLLOWON', fake.paragraph(nb_sentences=1), fake.paragraph(nb_sentences=1), '[today+' + str(random.randint(0, 10))  +']','1', '', '', fake.paragraph(nb_sentences=1)
                            ] # CANNOT USE D8352 = 'FOLLOWON' because it then is depenent on other items in T222.W

        #gererate test case sequence in Excel file       
        for i in range(TEST_CASE_LENGTH):
            # write transaction number to column 5
            # if looped many times, repeat every TEST_CASE_LENGTH rows
            # put the transaction number in sheet Test Case Sequence, column E - Test Step ref
            ws11.cell(row=i+4+(a*TEST_CASE_LENGTH), column=5).value = TEST_CASE_SEQUENCE[i]
            if ws11.cell(row=i+4+(a*TEST_CASE_LENGTH), column=5).value == 'T501.R':
                ws11.cell(row=i+4+(a*TEST_CASE_LENGTH), column=5).fill = PatternFill(start_color="CECE0B", fill_type = "solid")
            # build file name based on transactions chain. i.e. T321W_T201W_T322W....
            # new_filename = new_filename + TEST_CASE_SEQUENCE[i] + '_'
            # if transaction has .R in the name, it is MOSLTEST-R as requestor
            # put the transaction Source Org ID in sheet Test Case Sequence, column C - Source ID
            if TEST_CASE_SEQUENCE[i][-1] == 'R':
                ws11.cell(row=i+4+(a*TEST_CASE_LENGTH), column=3).value = RETAILER
            else:
                ws11.cell(row=i+4+(a*TEST_CASE_LENGTH), column=3).value = WHOLESALER
            # then in second sheet 'Test case data' depending on the transaction, insert respctive data items
            match TEST_CASE_SEQUENCE[i]:
                case 'T501.R':
                    ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=5).fill = PatternFill(start_color="DFDF00", fill_type = "solid")
                    for cols in range(7,21): # color basic request items
                        ws12.cell(row=4+(3*i)+(3*a*TEST_CASE_LENGTH), column=cols).fill = PatternFill(start_color="99FFCC", fill_type = "solid")
                    for cols in range(21,44): # color customer items
                        ws12.cell(row=4+(3*i)+(3*a*TEST_CASE_LENGTH), column=cols).fill = PatternFill(start_color="FFE5CC", fill_type = "solid")

                    for k in range(len(T501R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T501R_data_items[k]
                case 'T201.W':
                    for k in range(len(T201W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T201W_data_items[k]
                case 'T202.W':
                    for k in range(len(T202W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T202W_data_items[k]
                case 'T203.W':
                    for k in range(len(T203W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T203W_data_items[k]
                case 'T204.R':
                    for k in range(len(T204R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T204R_data_items[k]
                case 'T207.R':
                    for k in range(len(T207R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T207R_data_items[k]
                case 'T207.W':
                    for k in range(len(T207W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T207W_data_items[k]
                case 'T208.R':
                    for k in range(len(T208R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T208R_data_items[k]
                case 'T210.R':
                    for k in range(len(T210R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T210R_data_items[k]
                case 'T211.R':
                    for k in range(len(T211R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T211R_data_items[k]                  
                case 'T213.W':
                    for k in range(len(T213W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T213W_data_items[k]                   
                case 'T222.W':
                    for k in range(len(T222W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T222W_data_items[k]

    # remove lsast underscore from file name: T351R_T201W_T352W_                          
    # new_filename = new_filename[:-1]
    # and save it with an xlsx extension
    new_filename = '_'.join(TEST_CASE_SEQUENCE)
    
    test_cases_folder = working_dir + 'TEST_CASES'
    if not os.path.exists(test_cases_folder):
        os.makedirs(test_cases_folder)

    if loop_times > 1:
        suffix = '_X' + str(loop_times)

    wb1.save(filename = test_cases_folder + '\\' + new_filename.replace('.','') + suffix + '.xlsx')

# loop_times repeats test case sequence in the excel file
#max_loop = int (100/TEST_CASE_LENGTH)
max_loop = 5
# if you want to enter manually number of rows, do not exceed 100 - 100 / TEST_CASE_LENGTH shouldn't be bigger then max_loop!
generate_test_case_F5R(max_loop)