import openpyxl as xl
import os
from openpyxl.styles import numbers

# list files in current directory
mylist = os.listdir("C:\\Users\\tomasz.skoczylas\\Downloads\\11\\")
# remove the template file from the list, so it is not used as source file
mylist.remove('Bilaterals 1.1 master.xlsx')

for f in mylist:

    # opening the source excel file with first two worksheets
    filename1 = "C:\\Users\\tomasz.skoczylas\\Downloads\\11\\" + f
    wb1 = xl.load_workbook(filename1)
    ws11 = wb1.worksheets[0]
    ws12 = wb1.worksheets[1]

    # opening the destination excel file
    filename2 = "C:\\Users\\tomasz.skoczylas\\Downloads\\11\\Bilaterals 1.1 master.xlsx"
    wb2 = xl.load_workbook(filename2)
    ws21 = wb2.worksheets[0]
    ws22 = wb2.worksheets[1]
    #finding number of transactions
    trx_no = 0
    for t in range (4,24):
        if ws11.cell(row=t, column=3).value  != '-':
            trx_no += 1
    #print('File ' + f + ' - number of transactions: '+ str(trx_no))

    #copy first worksheet - Test Case Sequence
    for i in range(4, 4+trx_no):
        for j in range(3, 6):
            c = ws11.cell(row=i, column=j)
            ws21.cell(row=i, column=j).value = c.value
    # add info about version
    ws21.cell(
        row=4, column=7).value = 'created with 1.1 excel template'
    #copy first worksheet - Test Case Sequence

    #copy second worksheet - Test Data

    if ws11.cell(row=4, column=5).value  == 'T351.R':
        #copy first part of T351.R
        for k in range(7, 35):
            c1 = ws12.cell(row=6, column=k)
            ws22.cell(row=6, column=k).value = c1.value
        #related request ORID is empty / database was flushed
        ws22.cell(row=6, column=9).value =''
        #new data item D8330 Address as in CMOS
        ws22.cell(row=6, column=35).value = '1'
        #copy second part of T351.R - shift one column in new nworksheet
        for n in range(35, 59):
            c2 = ws12.cell(row=6, column=n)
            ws22.cell(row=6, column=n+1).value = c2.value
        ws22.cell(row=6, column=58).number_format = numbers.FORMAT_GENERAL
        ws22.cell(row=6, column=58).value = '2021-11-01'
        #print('T351.R')

    elif ws11.cell(row=4, column=5).value  == 'T351.W':
        #copy first part of T351.W
        for k in range(7, 13):
            c1 = ws12.cell(row=6, column=k)
            ws22.cell(row=6, column=k).value = c1.value
        #related request ORID is empty / database was flushed
        ws22.cell(row=6, column=8).value =''
        #new data item D8330 Address as in CMOS
        ws22.cell(row=6, column=13).value = '1'
        #copy second part of T351.W - shift one column in new nworksheet
        for n in range(13, 37):
            c2 = ws12.cell(row=6, column=n)
            ws22.cell(row=6, column=n+1).value = c2.value
        ws22.cell(row=6, column=36).number_format = numbers.FORMAT_GENERAL
        ws22.cell(row=6, column=36).value = '2021-11-01'
        #print('T351.W')

    else:
        for k in range(7, 58):
            c1 = ws12.cell(row=6, column=k)
            ws22.cell(row=6, column=k).value = c1.value
        #print ('copy 1-1')
    
    #copy rest of the transactions if they exist
    if trx_no > 1:
            #copy row number based on transactions number
            for k in range(9, 4+(trx_no*3)):
                #copy columns from 7 to 40
                for l in range(7, 40):
                    d = ws12.cell(row=k, column=l)
                    ws22.cell(row=k, column=l).value = d.value


    wb2.save(str(filename1.replace(".xlsx", "_new.xlsx")))