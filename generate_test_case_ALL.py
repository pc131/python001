import openpyxl as xl
import os
from datetime import datetime
import random
import string
import names
import sys
from random_address import real_random_address
from faker import Faker
from faker_biology.mol_biol import Enzyme

#working_dir = "C:\\Users\\skocz\\Downloads\\11\\"
working_dir = "C:\\Users\\tomasz.skoczylas\\Downloads\\11\\"
filename = working_dir + "Bilaterals 1.4.0.0 master.xlsx"
wb1 = xl.load_workbook(filename)
ws11 = wb1.worksheets[0]
ws12 = wb1.worksheets[1]


TEST_CASE_TRANSACTIONS = ['T355.W','T201.W','T217.W']
SPIDS_METERS = {'3019178819W13':('KENT','000000000000189985'),'3019178827W10':('KENT','4A059235'),'3019178843W15':('KENT','4A026515'),'3019178851W12':('ELSTER','000000000008081344'),'301917886XW1X':('ARAD','152026121'),'3019178991W12':('KENT','000000000000221909'),'3019179017W10':('KENT','4A024157'),'3019179033W15':('KENT','2T021233'),'3019179130W16':('KENT','90P169156'),'3019179149W13':('KENT','3T042349'),'3019179173W15':('ARAD','09044840'),'3019179289W13':('KENT','AE188861'),'301917936XW1X':('KENT','4A023209'),'3019179386W14':('ARAD','16732739'),'3019179416W14':('ARAD','8004476'),'3019179483W15':('AMR','8569517'),'3019179564W11':('ARAD','9103911'),'3019179653W15':('KENT','000000000083235847'),'3019024854W11':('ARAD','9097878'),'3019025117W10':('SW_METER','9M129866'),'3019025125W18':('KENT','2M034061'),'3019025230W16':('KENT','3M310071'),'3019025486W14':('KENT','4M093717'),'3019025532W19':('ARAD','000000000008011122'),'3019025648W17':('ARAD','000000000008596289'),'3019025648W17':('ARAD','000000000008385953'),'3019025737W10':('KENT','4M058479'),'3019026067W10':('KENT','1M079127'),'301902644XW1X':('KENT','000000000090248995'),'3019027314W11':('ARAD','000000000008110959'),'3019027322W19':('KENT','1M202328'),'301902739XW1X':('KENT','000000000087080861'),'3019027691W12':('KENT','000000000085077189'),'3019027780W16':('SW_METER','7M472116'),'3019028310W16':('KENT','94A003592'),'301902840XW1X':('KENT','4A044797'),'3019028574W11':('KENT','5M049422'),'3019179726W14':('KENT','4T024221'),'3019179750W16':('ELSTER','7A165214'),'3019179793W15':('KENT','000000000089116779'),'3019179858W17':('ELSTER','000000000008064363'),'3019179874W11':('KENT','000000000000191037'),'3019179890W16':('ARAD','18AI0248'),'3019179912W19':('ARAD','000000000008295070'),'3019180058W17':('KENT','000000000085179813'),'3019180244W11':('ARAD','09135365'),'3019180309W13':('KENT','000000000082099104'),'3019180368W17':('SCHLUMBERGER','96AQ151103'),'3019180384W11':('ARAD','9135944'),'3019180449W13':('KENT','000000000001489140'),'3019180465W18':('ARAD','9114869'),'3019180473W15':('NEPTUNE_MEASUREMENT','000000000000221940'),'3019180856W14':('KENT','4T007899'),'3019180902W19':('AQUADIS','96AL002538'),'3019180929W13':('KENT','4T019536'),'3019181127W10':('KENT','1M068999'),'3019181240W16':('KENT','4T033566'),'3019181259W13':('ARAD','000000000008232146'),'3019181313W15':('NEPTUNE_MEASUREMENT','1T024790'),'3019181372W19':('NEPTUNE_MEASUREMENT','93M054980'),'3019181429W13':('SW_METER','9A072507'),'3019181518W17':('KENT','4T025116'),'3019181682W19':('AQUADIS','97PC602000'),'3019181763W15':('ARAD','000000000008280228'),'3019181798W17':('ARAD','8174889'),'3019181860W16':('KENT','4T018841'),'3019181895W18':('FROST','AL022639'),'3019181917W10':('KENT','000000000082089093'),'3019181925W18':('KENT','4T012674'),'3019182085W18':('KENT','3M301521'),'3019182166W14':('ELSTER','000000000008496706'),'3019182174W11':('ARAD','9164269'),'3019182247W10':('SCHLUMBERGER','97AQ268415'),'3019182484W11':('SCHLUMBERGER','96AQ123266'),'3019182506W14':('KENT','2M000061'),'3019182980W16':('KENT','95M186330'),'3019183081W12':('KENT','3M335173'),'301918309XW1X':('ARAD','000000000008017074'),'3019183111W12':('KENT','91M126435'),'3019183227W10':('KENT','6T007525'),'3019183286W14':('KENT','000000000085179921'),'3019183308W17':('KENT','000000000000023101'),'3019183316W14':('ELSTER','12A10363'),'3019183324W11':('KENT','93A023199'),'3019183340W16':('KENT','000000000087164732'),'301918343XW1X':('ELSTER','000000000008093998'),'3019029031W12':('KENT','4T007325'),'3019029058W17':('ELSTER','1M219584'),'3019029147W10':('ARAD','9158366'),'3019029368W17':('ARAD','000000000008378520'),'3019029449W13':('ARAD','000000000008174844'),'3019029457W10':('KENT','93M213479'),'3019029546W14':('KENT','95A138505'),'301903020XW1X':('KENT','4M097305'),'3019030803W15':('KENT','1A161821'),'3019031141W12':('KENT','4M117380'),'3019031214W11':('AQUADIS','98PB300065'),'301903129XW1X':('ELSTER','4M093252'),'3019031559W13':('KENT','95A521326'),'3019031567W10':('KENT','93M154072'),'3019031575W18':('ELSTER','000000000008094716'),'3019031591W12':('ELSTER','000000000008999775'),'3019031621W12':('ARAD','000000000008270804'),'3019031850W16':('KENT','4M097140'),'3019032008W17':('KENT','2M224449'),'3019032377W10':('KENT','4T007843'),'3019032474W11':('SW_METER','4A061558'),'3019183553W15':('SCHLUMBERGER','99WPE05580'),'3019183642W19':('ARAD','000000000008963442'),'3019183693W15':('KENT','94M073927'),'3019183723W15':('KENT','000000000082089069'),'3019183790W16':('KENT','95A517029'),'3019183820W16':('ELSTER','000000000008103182'),'3019183855W18':('KENT','4T016561'),'3019183898W17':('KENT','4A023537'),'3019183952W19':('NA','9045596'),'3019184045W18':('KENT','000000000083103119'),'3019184061W12':('KENT','5A026406'),'301918410XW1X':('KENT','AF518194'),'3019184320W16':('KENT','4M201816'),'3019184363W15':('KENT','4M201366'),'3019184371W12':('KENT','4M117218'),'3019184495W18':('KENT','000000000085134566'),'3019184509W13':('SCHLUMBERGER','96AQ183368'),'3019184525W18':('KENT','94M004196'),'3019184770W16':('ARAD','000000000009050719'),'3019184878W17':('ELSTER','000000000003426633'),'301918522XW1X':('KENT','95M246039'),'3019185297W10':('SCHLUMBERGER','95A131611'),'3019185319W13':('NEPTUNE_MEASUREMENT','000000000091506905'),'3019185661W12':('ARAD','000000000008540387'),'3019185688W17':('ARAD','8219393'),'301903258XW1X':('KENT','93A556059'),'3019032601W12':('SW_METER','8M223846'),'301903261XW1X':('SW_METER','8M473776'),'3019032628W17':('KENT','91P033387'),'301903275XW1X':('ELSTER','000000000012207830'),'3019032938W17':('ARAD','9103768'),'3019032946W14':('ARAD','19AI0097'),'3019032954W11':('ELSTER','9097938'),'3019032997W10':('KENT','PC502819'),'3019033004W11':('KENT','AE183174'),'3019033012W19':('KENT','79011716'),'3019033020W16':('KENT','000000000073225966'),'3019033039W13':('KENT','97A080026'),'3019033047W10':('ARAD','08013499'),'301903308XW1X':('ARAD','000000000008567195'),'3019033101W12':('KENT','AM014229'),'301903311XW1X':('KENT','529791'),'3019033128W17':('KENT','000000000006259797'),'3019033136W14':('ELSTER','10AC017906'),'3019033144W11':('KENT','000000000082099364'),'3019033152W19':('KENT','4A054549'),'3019033160W16':('NEPTUNE_MEASUREMENT','000000000000321187'),'3019033195W18':('ARAD','000000000008584440'),'3019185696W14':('KENT','088174268'),'301918570XW1X':('KENT','AG610576'),'3019185815W18':('NEPTUNE_MEASUREMENT','99AL005090'),'3019185963W15':('KENT','000000000089582490'),'3019186064W11':('ARAD','000000000008360573'),'3019186080W16':('SW_METER','8M086205'),'301918620XW1X':('KENT','000000000073339733'),'3019186218W17':('KENT','4A045367'),'3019186226W14':('ARAD','000000000008581723'),'3019186234W11':('KENT','93A021395'),'3019186250W16':('ELSTER','000000000008217650'),'3019186285W18':('KENT','3M266548'),'3019186315W18':('KENT','93M126286'),'3019186323W15':('KENT','000000000078144637'),'301918634XW1X':('KENT','PC502050'),'3019186366W14':('KENT','000000000078144630'),'3019186374W11':('KENT','000000000079113691'),'3019186412W19':('KENT','000000000080079221'),'3019186439W13':('ARAD','000000000008119893'),'3019186552W19':('KENT','4T012332'),'3019186560W16':('KENT','4A061265'),'3019186587W10':('KENT','4A061268'),'3019186609W13':('KENT','000000000000610563'),'3019186633W15':('ARAD','16286125'),'3019186668W17':('KENT','000000000073339734'),'3019186676W14':('SW_METER','9M174577'),'3019186765W18':('SOCAM','000000000004748629'),'3019186781W12':('ELSTER','000000000008036473'),'3019187192W19':('ELSTER','8M001142'),'301918732XW1X':('SW_METER','7A128306'),'3019187400W16':('KENT','4A029940'),'3019187516W14':('KENT','90P194329'),'3019187524W11':('KENT','4M093692'),'3019187613W15':('ELSTER','000000000008951107'),'3019187729W13':('SW_METER','8A026502'),'301918777XW1X':('KENT','4A061242'),'3019187850W16':('KENT','93A579245'),'3019187907W10':('KENT','000000000073279189'),'3019188059W13':('ELSTER','000000000008092307'),'3019188105W18':('KENT','4A022348'),'3019188172W19':('ARAD','9116984'),'3019188202W19':('KENT','AG610567'),'3019188245W18':('ABB','G/09289/1/1'),'3019188253W15':('KENT','4A072078'),'3019188296W14':('ARAD','8189888'),'3019188334W11':('KENT','4T013708'),'3019188385W18':('KENT','92M047166'),'3019188458W17':('SW_METER','8M088276'),'3019188466W14':('KENT','M210093'),'3019188490W16':('KENT','6A048505'),'3019188512W19':('KENT','000000000000405089'),'3019033276W14':('ARAD','9089368'),'3019033306W14':('KENT','4T009453'),'3019033322W19':('ELSTER','000000000009080130'),'3019033330W16':('ELSTER','000000000008288295'),'3019033373W15':('ARAD','9108738'),'3019033381W12':('KENT','000000000073260854'),'3019033411W12':('SW_METER','9M129985'),'301903342XW1X':('KENT','AL034255'),'3019033438W17':('NEPTUNE_MEASUREMENT','368175'),'3019033462W19':('KENT','4A052252'),'3019033500W16':('KENT','4T015998'),'3019033535W18':('KENT','4A022358'),'3019033578W17':('KENT','86027961'),'3019033594W11':('KENT','4A022325'),'3019033624W11':('KENT','000000000073261507'),'3019033632W19':('KENT','000000000082099358'),'3019033659W13':('ARAD','9124697'),'3019033667W10':('KENT','000000000088457019'),'3019033721W12':('KENT','000000000000638354'),'301903373XW1X':('KENT','3M060812'),'3019033829W13':('KENT','4A023413'),'3019033837W10':('SW_METER','8T000006'),'3019033845W18':('KENT','3A246396'),'3019033853W15':('KENT','000000000008507331'),'3019033888W17':('ARAD','9120052'),'3019033926W14':('SW_METER','14AI0081'),'3019123186W14':('KENT','87148248'),'3019123194W11':('ELSTER','000000000008360024'),'3019123275W18':('ARAD','000000000008536005'),'3019123364W11':('SCHLUMBERGER','96AQ133204'),'3019123453W15':('ELSTER','10A033112'),'3019123518W17':('KENT','4T025198'),'3019123607W10':('SW_METER','9A081604'),'301912364XW1X':('ARAD','9417713'),'3019123682W19':('KENT','000000000072165998'),'3019123704W11':('KENT','3M158669'),'3019123917W10':('KENT','000000000078149353'),'3019123925W18':('KENT','000000000082096204'),'3019123933W15':('ARAD','8188202'),'3019123968W17':('ARAD','000000000009063868'),'3019124034W11':('KENT','000000000082089101'),'3019124050W16':('KENT','3M158614'),'3019124093W15':('ELSTER','8123500'),'3019124158W17':('KENT','5M045270'),'3019124174W11':('SCHLUMBERGER','99AQ448419'),'3019124190W16':('KENT','4T024226'),'3019124263W15':('SCHLUMBERGER','96AQ138070'),'3019124271W12':('KENT','3M073842'),'3019124328W17':('KENT','4T008475'),'3019124409W13':('ELSTER','000000000008189180'),'3019124417W10':('ELSTER','000000000008034615'),'3019124425W18':('KENT','000000000082089158'),'3019124433W15':('KENT','4T009051'),'3019391148W17':('ARAD','000000000008075195'),'3019391407W10':('KENT','4M097120'),'3019391849W13':('KENT','4T012147'),'301939189XW1X':('SCHLUMBERGER','99PB401460'),'3019391938W17':('KENT','5M370619'),'3019392500W16':('ARAD','9118075'),'301939273XW1X':('ARAD','09089691'),'301939287XW1X':('KENT','4M060742'),'3019393213W15':('SEVERN_TRENT_(ELECTRONIC)','000000000000000072'),'3019393221W12':('KENT','5M006818'),'301939323XW1X':('ARAD','09141412'),'3019393280W16':('KENT','88073486'),'3019393396W14':('KENT','4T014731'),'3019393639W13':('KENT','2M002958'),'3019394007W10':('KENT','4A072064'),'3019394090W16':('KENT','4A219373'),'3019394139W13':('KENT','4M261957'),'3019394147W10':('KENT','5M020001'),'3019394171W12':('KENT','4A071623'),'3019394309W13':('KENT','4A052357'),'3019394481W12':('KENT','4T041981'),'3019394716W14':('KENT','6M145886'),'3019394724W11':('KENT','2M074007'),'3019394759W13':('SCHLUMBERGER','99AQ503224'),'3019394805W18':('KENT','6M294789'),'3019394902W19':('KENT','4A052265'),'3019394953W15':('KENT','4T031858'),'3019394961W12':('KENT','4A052354'),'3019395267W10':('KENT','90P57308'),'3019395291W12':('KENT','1M243814'),'3019395321W12':('ARAD','09416996'),'3019395364W11':('KENT','4A023500'),'3019395429W13':('KENT','4M346891'),'301939550XW1X':('ARAD','16732862'),'3019395550W16':('ARAD','9104838'),'3019395593W15':('SCHLUMBERGER','AL022619'),'3019395615W18':('ARAD','9147350'),'3019395852W19':('KENT','9106312'),'3019395976W14':('ARAD','9161215'),'3019396026W14':('KENT','2A156701'),'3019396115W18':('KENT','000000000000525071')}
#2WSL_2RTL
#SPIDS_METERS = {'3200643021W12':('ELSTER','08019965'),'3200647175W18':('ELSTER','10W10270746134'),'320065323XW1X':('KENT','92A588101'),'3200663693W15':('ELSTER','06M294533'),'3200675837W10':('ARAD','9129186'),'3200681691W12':('ELSTER','4A064322'),'3200719893W15':('ARAD','9132435'),'3200720123W15':('ARAD','9140801'),'320072031XW1X':('ARAD','9139171'),'3200727020W16':('ARAD','9131614'),'3200751924W11':('ARAD','16287021'),'320076144XW1X':('ARAD','9149113'),'3200767723W15':('ELSTER','AM076399'),'3200768703W15':('Elster','4T006734'),'3200772247W10':('SW_METER','7T020818'),'3200772344W11':('KENT','2A198645'),'3200795069W13':('ARAD','9157122'),'3200795700W16':('ARAD','9153721'),'3200795824W11':('ARAD','16284757'),'3200795956W14':('ARAD','16284749'),'3200796227W10':('ARAD','16735108'),'320080758XW1X':('ARAD','9153952'),'3200808098W17':('ARAD','16287218'),'320081974XW1X':('ELSTER','8994469'),'320082252XW1X':('ARAD','9157217'),'3200832975W18':('ELSTER','9045272'),'3200854618W17':('ARAD','9134010'),'3200858052W19':('TAGUS','13A10338'),'3200882530W16':('ARAD','9163543'),'3200887478W17':('ARAD','09150547'),'320090190XW1X':('ARAD','9160017'),'3200908998W17':('ARAD','17A0001122'),'3200914556W14':('ARAD','9317073'),'3200914645W18':('ARAD','9412020'),'3200921196W14':('SW_METER','8M097098'),'3200924446W14':('ARAD','9164296'),'3200925795W18':('ARAD','9315201'),'3200930713W15':('ARAD','9156780'),'3200937726W14':('ARAD','9511653'),'3200946997W10':('Elster','82089120'),'3200951575W18':('ARAD','9416836'),'3200966254W11':('ARAD','9161998'),'3201012467W10':('ARAD','9133188'),'3201014737W10':('ARAD','9155967'),'3201033294W11':('ARAD','8554218'),'320104721XW1X':('ARAD','9416445'),'3201047260W16':('ARAD','9414050'),'3201050164W11':('ARAD','9510449'),'3201052868W17':('ARAD','20MS000335'),'3201061018W17':('ARAD','08966422')}

RETAILER = 'MOSLTEST-R'
WHOLESALER = 'MOSLTEST-W'
T216_URL = 'https://moservicesdev.mosl.co.uk/test/attachments/87ffc85e-ebd5-461c-99d6-2ac3eef43f7c'

POSTCODES = ['B1 1HQ', 'BN88 1AH', 'BS98 1TL', 'BX1 1LT', 'BX2 1LB', 'BX3 2BB', 'BX4 7SB', 'BX5 5AT', 'CF10 1BH', 'CF99 1NA', 'CO4 3SQ', 'CV4 8UW', 'CV35 0DB', 'E14 5EY', 'DA1 1RT', 'DE99 3GG', 'DE55 4SW', 'DH98 1BT', 'DH99 1NS', 'E14 5HQ', 'E14 5JP', 'E16 1XL', 'E20 2AQ', 'E20 2BB', 'E20 2ST', 'E20 3BS', 'E20 3EL', 'E20 3ET', 'E20 3HB', 'E20 3HY', 'E98 1SN', 'E98 1ST', 'E98 1TT', 'EC2N 2DB', 'EC4Y 0HQ', 'EH12 1HQ', 'EH99 1SP', 'G58 1SB', 'GIR 0AA', 'IV21 2LR', 'L30 4GB', 'LS98 1FD', 'M50 2BH', 'M50 2QH', 'N1 9G', 'N81 1ER', 'NE1 4ST', 'NG80 1EH', 'NG80 1LH', 'NG80 1RH', 'NG80 1TH', 'PH1 5RB', 'PH1 2SJ', 'S2 4SU	', 'S6 1SW', 'S14 7UP', 'SE1 0NE', 'SE1 8UJ', 'SM6 0HB', 'SN38 1NW', 'SR5 1SU', 'SW1A 0AA', 'SW1A 0PW', 'SW1A 1AA', 'SW1A 2AA', 'SW1A 2AB', 'SW1H 0TL', 'SW1P 3EU', 'SW1W 0DT', 'SW11 7US', 'SW19 5AE', 'TW8 9GS', 'W1A 1AA', 'W1D 4FA', 'W1N 4DJ', 'W1T 1FB']

### MAKE SURE THERE IS THE SAME NUMBER OF ELEMENTS IN PROCESSES AND PROC_NAMES DICTIONARIES/TUPLES !!!!!!!!!!!!!!
PROCESSES = ['B1R', 'B3R', 'B3W', 'B5R', 'B5W', 'B7R', 'C1R', 'C1W', 'F4R', 'F4W', 'F5R', 'F5W', 'G1R', 'G1W', 'G2AR', 'G2AW', 'G2BR', 'G2BW']

PROC_NAMES = {'B1R':'Request Meter Install Work', 'B3R':'Request Meter Accuracy Test', 'B3W':'Request Meter Accuracy Test', 'B5R':'Request Meter Repair Replacement Work', 'B5W':'Request Meter Repair Replacement Work', 'B7R':'Request Meter Change', 'C1R':'Request Meter And Supply Arrangement Verification', 'C1W':'Request Meter And Supply Arrangement Verification', 'F4R':'Submit Non-Household Customer Enquiry', 'F4W':'Submit Non-Household Customer Enquiry', 'F5R':'Submit Non-Household Customer Complaint', 'F5W':'Submit Non-Household Customer Complaint', 'G1R':'Submit Non-Household Customer TE Enquiry', 'G1W':'Submit Non-Household Customer TE Enquiry', 'G2AR':'Submit TE Consent Application With SPID', 'G2AW':'Submit TE Consent Application With SPID', 'G2BR':'Submit TE Consent Application Without SPID', 'G2BW':'Submit TE Consent Application Without SPID'}

TRANSACTION_NAMES = {'T201.W':'Accept Service Request', 'T202.W':'Reject Service Request', 'T203.W':'Request For Additional Information', 'T204.R':'Provide Additional Information', 'T205.W':'Update Site Visit Date', 'T206.W':'Update Site Visit Failure', 'T207.R':'Submit Trading Party Comments', 'T207.W':'Submit Trading Party Comments', 'T208.R':'Close Service Request', 'T210.R':'Resubmit Service Request', 'T211.R':'Cancel Service Request', 'T211.W':'Cancel Service Request', 'T212.W':'Visit Complete And Preparing Plan', 'T213.W':'Start Service Request Deferral', 'T214.W':'End Service Request Deferral', 'T215.R':'Provide Attachment', 'T215.W':'Provide Attachment', 'T216.R':'Request Attachment', 'T216.W':'Request Attachment', 'T217.W':'Request For Customer Details and Additional Information', 'T218.R':'Provide Customer Details and Additional Information', 'T220.W':'Provide Quote For Non Standard Activity', 'T221.R':'Accept Quote For Non Standard Activity', 'T222.W':'Advise Service Request Complete', 'T223.W':'Advise Meter Work Completion', 'T224.W':'Advise Process Delay', 'T321.R':'Request Meter And Supply Arrangement Verification', 'T321.W':'Request Meter And Supply Arrangement Verification', 'T322.W':'Update Corrections Complete for C1', 'T323.W':'Propose Corrections Plan for C1', 'T324.R':'Agree Proposed Corrections Plan for C1', 'T325.R':'Dispute Proposed Corrections Plan for C1', 'T351.R':'Request Meter Repair Replacement Work', 'T351.W':'Request Meter Repair Replacement Work', 'T352.W':'Advise Meter Repair Replacement Work Completion', 'T353.R':'Request Meter Install Work', 'T355.R':'Request Meter Accuracy Test', 'T355.W':'Request Meter Accuracy Test', 'T356.W':'Advise Meter Accuracy Test Complete', 'T357.W':'Awaiting Meter Accuracy Test', 'T365.R':'Request Meter Change', 'T501.R':'Submit Non Household Customer Complaint', 'T501.W':'Submit Non Household Customer Complaint', 'T505.R':'Submit Non Household Customer Enquiry', 'T505.W':'Submit Non Household Customer Enquiry', 'T551.R':'Submit Non Household Customer TE Enquiry', 'T551.W':'Submit Non Household Customer TE Enquiry', 'T555.R':'Submit TE Consent Application With SPID', 'T555.W':'Submit TE Consent Application With SPID', 'T556.R':'Submit TE Consent Application Without SPID', 'T556.W':'Submit TE Consent Application Without SPID', 'T557.W':'Advise TE Consent Application Outcome'}

C1R_TRANSACTIONS = ['T321.R', 'T201.W', 'T202.W', 'T203.W', 'T204.R', 'T205.W', 'T206.W', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T212.W', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T322.W', 'T323.W', 'T324.R', 'T325.R']
C1R_T201W_allowed = ['T203.W', 'T205.W', 'T322.W', 'T323.W']
C1R_T202W_allowed = ['T210.R']
C1R_T203W_allowed = ['T204.R']
C1R_T204R_allowed = ['T203.W', 'T205.W', 'T322.W', 'T323.W']
C1R_T205W_allowed = ['T206.W', 'T212.W', 'T322.W', 'T323.W']
C1R_T206W_allowed = ['T203.W', 'T205.W']
C1R_T210R_allowed = ['T201.W', 'T202.W']
C1R_T212W_allowed = ['T203.W', 'T323.W']
C1R_T321R_allowed = ['T201.W', 'T202.W']
C1R_T322W_allowed = ['T208.R', 'T210.R']
C1R_T323W_allowed = ['T324.R', 'T325.R']
C1R_T324R_allowed = ['T203.W', 'T205.W', 'T322.W']
C1R_T325R_allowed = ['T203.W', 'T323.W']

C1W_TRANSACTIONS = ['T321.W', 'T201.W', 'T202.W', 'T217.W', 'T218.R', 'T205.W', 'T206.W', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T212.W', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T322.W', 'T323.W', 'T324.R', 'T325.R']
C1W_T201W_allowed = ['T217.W', 'T205.W', 'T322.W', 'T323.W']
C1W_T202W_allowed = ['T210.R']
C1W_T217W_allowed = ['T218.R']
C1W_T218R_allowed = ['T217.W', 'T205.W', 'T322.W', 'T323.W']
C1W_T205W_allowed = ['T206.W', 'T212.W', 'T322.W', 'T323.W']
C1W_T206W_allowed = ['T217.W', 'T205.W']
C1W_T210R_allowed = ['T201.W', 'T202.W']
C1W_T212W_allowed = ['T217.W', 'T323.W']
C1W_T321W_allowed = ['T201.W']
C1W_T322W_allowed = ['T208.R', 'T210.R']
C1W_T323W_allowed = ['T324.R', 'T325.R']
C1W_T324R_allowed = ['T217.W', 'T205.W', 'T322.W']
C1W_T325R_allowed = ['T217.W', 'T323.W']

B5R_TRANSACTIONS = ['T351.R', 'T201.W', 'T202.W', 'T203.W', 'T204.R', 'T205.W', 'T206.W', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T224.W', 'T352.W']
B5R_T201W_allowed = ['T203.W', 'T205.W', 'T224.W', 'T352.W']
B5R_T202W_allowed = ['T210.R']
B5R_T203W_allowed = ['T204.R']
B5R_T204R_allowed = ['T203.W', 'T205.W', 'T224.W', 'T352.W']
B5R_T205W_allowed = ['T206.W', 'T224.W', 'T352.W']
B5R_T206W_allowed = ['T203.W', 'T205.W']
B5R_T210R_allowed = ['T201.W', 'T202.W']
B5R_T224W_allowed = ['T203.W', 'T205.W', 'T352.W']
B5R_T351R_allowed = ['T201.W', 'T202.W']
B5R_T352W_allowed = ['T208.R', 'T210.R']

B5W_TRANSACTIONS = ['T351.W', 'T201.W', 'T202.W', 'T217.W', 'T218.R', 'T205.W', 'T206.W', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T224.W', 'T352.W']
B5W_T201W_allowed = ['T217.W', 'T205.W', 'T224.W', 'T352.W']
B5W_T202W_allowed = ['T210.R']
B5W_T217W_allowed = ['T218.R']
B5W_T218R_allowed = ['T217.W', 'T205.W', 'T224.W', 'T352.W']
B5W_T205W_allowed = ['T206.W', 'T224.W', 'T352.W']
B5W_T206W_allowed = ['T217.W', 'T205.W']
B5W_T210R_allowed = ['T201.W', 'T202.W']
B5W_T224W_allowed = ['T217.W', 'T205.W', 'T352.W']
B5W_T351W_allowed = ['T201.W']
B5W_T352W_allowed = ['T208.R', 'T210.R']

B1R_TRANSACTIONS = ['T353.R', 'T201.W', 'T202.W', 'T203.W', 'T204.R', 'T205.W', 'T206.W', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T220.W', 'T221.R', 'T224.W', 'T223.W']
B1R_T353R_allowed = ['T201.W', 'T202.W']
B1R_T201W_allowed = ['T203.W', 'T205.W', 'T220.W', 'T224.W', 'T223.W']
B1R_T202W_allowed = ['T210.R']
B1R_T203W_allowed = ['T204.R']
B1R_T204R_allowed = ['T203.W', 'T205.W', 'T220.W', 'T224.W', 'T223.W']
B1R_T205W_allowed = ['T206.W', 'T220.W', 'T224.W', 'T223.W']
B1R_T206W_allowed = ['T203.W', 'T205.W']
B1R_T210R_allowed = ['T201.W', 'T202.W']
B1R_T220W_allowed = ['T210.R', 'T221.R']
B1R_T221R_allowed = ['T205.W', 'T224.W', 'T223.W']
B1R_T224W_allowed = ['T203.W', 'T205.W', 'T220.W', 'T223.W']
B1R_T223W_allowed = ['T208.R', 'T210.R']

B3R_TRANSACTIONS = ['T355.R', 'T201.W', 'T202.W', 'T203.W', 'T204.R', 'T205.W', 'T206.W', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T220.W', 'T221.R', 'T224.W', 'T356.W', 'T357.W']
B3R_T355R_allowed = ['T201.W', 'T202.W']
B3R_T201W_allowed = ['T203.W', 'T205.W', 'T220.W', 'T224.W', 'T356.W']
B3R_T202W_allowed = ['T210.R']
B3R_T203W_allowed = ['T204.R']
B3R_T204R_allowed = ['T203.W', 'T205.W', 'T220.W', 'T224.W', 'T356.W']
B3R_T205W_allowed = ['T206.W', 'T220.W', 'T224.W', 'T357.W', 'T356.W']
B3R_T206W_allowed = ['T203.W', 'T205.W']
B3R_T210R_allowed = ['T201.W', 'T202.W']
B3R_T220W_allowed = ['T210.R', 'T221.R']
B3R_T221R_allowed = ['T205.W', 'T224.W', 'T356.W']
B3R_T224W_allowed = ['T203.W', 'T205.W', 'T220.W', 'T356.W']
B3R_T357W_allowed = ['T356.W']
B3R_T356W_allowed = ['T208.R', 'T210.R']

B3W_TRANSACTIONS = ['T355.W', 'T201.W', 'T202.W', 'T217.W', 'T218.R', 'T205.W', 'T206.W', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T220.W', 'T221.R', 'T224.W', 'T356.W', 'T357.W']
B3W_T355W_allowed = ['T201.W']
B3W_T201W_allowed = ['T217.W', 'T205.W', 'T220.W', 'T224.W', 'T356.W']
B3W_T202W_allowed = ['T210.R']
B3W_T217W_allowed = ['T218.R']
B3W_T218R_allowed = ['T217.W', 'T205.W', 'T220.W', 'T224.W', 'T356.W']
B3W_T205W_allowed = ['T206.W', 'T220.W', 'T224.W', 'T357.W', 'T356.W']
B3W_T206W_allowed = ['T217.W', 'T205.W']
B3W_T210R_allowed = ['T201.W', 'T202.W']
B3W_T220W_allowed = ['T210.R', 'T221.R']
B3W_T221R_allowed = ['T205.W', 'T224.W', 'T356.W']
B3W_T224W_allowed = ['T217.W', 'T205.W', 'T220.W', 'T356.W']
B3W_T357W_allowed = ['T356.W']
B3W_T356W_allowed = ['T208.R', 'T210.R']


B7R_TRANSACTIONS = ['T365.R', 'T201.W', 'T202.W', 'T203.W', 'T204.R', 'T205.W', 'T206.W', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T220.W', 'T221.R', 'T224.W', 'T223.W']
B7R_T365R_allowed = ['T201.W', 'T202.W']
B7R_T201W_allowed = ['T203.W', 'T205.W', 'T220.W', 'T224.W', 'T223.W']
B7R_T202W_allowed = ['T210.R']
B7R_T203W_allowed = ['T204.R']
B7R_T204R_allowed = ['T203.W', 'T205.W', 'T220.W', 'T224.W', 'T223.W']
B7R_T205W_allowed = ['T206.W', 'T220.W', 'T224.W', 'T223.W']
B7R_T206W_allowed = ['T203.W', 'T205.W']
B7R_T210R_allowed = ['T201.W', 'T202.W']
B7R_T220W_allowed = ['T210.R', 'T221.R']
B7R_T221R_allowed = ['T205.W', 'T224.W', 'T223.W']
B7R_T224W_allowed = ['T203.W', 'T205.W', 'T220.W', 'T223.W']
B7R_T223W_allowed = ['T208.R', 'T210.R']

F4R_TRANSACTIONS = ['T505.R', 'T201.W', 'T202.W', 'T203.W', 'T204.R', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T222.W']
F4R_T505R_allowed = ['T201.W', 'T202.W']
F4R_T201W_allowed = ['T203.W', 'T222.W']
F4R_T202W_allowed = ['T210.R']
F4R_T203W_allowed = ['T204.R']
F4R_T204R_allowed = ['T203.W', 'T222.W']
F4R_T210R_allowed = ['T201.W', 'T202.W']
F4R_T222W_allowed = ['T208.R', 'T210.R']

F4W_TRANSACTIONS = ['T505.W', 'T201.W', 'T202.W', 'T203.W', 'T204.R', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T222.W']
F4W_T505W_allowed = ['T201.W']
F4W_T201W_allowed = ['T217.W', 'T222.W']
F4W_T202W_allowed = ['T210.R']
F4W_T217W_allowed = ['T218.R']
F4W_T218R_allowed = ['T217.W', 'T222.W']
F4W_T210R_allowed = ['T201.W', 'T202.W']
F4W_T222W_allowed = ['T208.R', 'T210.R']

F5R_TRANSACTIONS = ['T501.R', 'T201.W', 'T202.W', 'T203.W', 'T204.R', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T222.W']
F5R_T501R_allowed = ['T201.W', 'T202.W']
F5R_T201W_allowed = ['T203.W', 'T222.W']
F5R_T202W_allowed = ['T210.R']
F5R_T203W_allowed = ['T204.R']
F5R_T204R_allowed = ['T203.W', 'T222.W']
F5R_T210R_allowed = ['T201.W', 'T202.W']
F5R_T222W_allowed = ['T208.R', 'T210.R']

F5W_TRANSACTIONS = ['T501.W', 'T201.W', 'T202.W', 'T203.W', 'T204.R', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T222.W']
F5W_T501W_allowed = ['T201.W']
F5W_T201W_allowed = ['T217.W', 'T222.W']
F5W_T202W_allowed = ['T210.R']
F5W_T217W_allowed = ['T218.R']
F5W_T218R_allowed = ['T217.W', 'T222.W']
F5W_T210R_allowed = ['T201.W', 'T202.W']
F5W_T222W_allowed = ['T208.R', 'T210.R']

G1R_TRANSACTIONS = ['T551.R', 'T201.W', 'T202.W', 'T203.W', 'T204.R', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T222.W']
G1R_T551R_allowed = ['T201.W', 'T202.W']
G1R_T201W_allowed = ['T203.W', 'T222.W']
G1R_T202W_allowed = ['T210.R']
G1R_T203W_allowed = ['T204.R']
G1R_T204R_allowed = ['T203.W', 'T222.W']
G1R_T210R_allowed = ['T201.W', 'T202.W']
G1R_T222W_allowed = ['T208.R', 'T210.R']

G1W_TRANSACTIONS = ['T551.W', 'T201.W', 'T202.W', 'T203.W', 'T204.R', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T222.W']
G1W_T551W_allowed = ['T201.W']
G1W_T201W_allowed = ['T217.W', 'T222.W']
G1W_T202W_allowed = ['T210.R']
G1W_T217W_allowed = ['T218.R']
G1W_T218R_allowed = ['T217.W', 'T222.W']
G1W_T210R_allowed = ['T201.W', 'T202.W']
G1W_T222W_allowed = ['T208.R', 'T210.R']

G2AR_TRANSACTIONS = ['T555.R', 'T201.W', 'T202.W', 'T203.W', 'T204.R', 'T205.W', 'T206.W','T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T224.W', 'T557.W']
G2AR_T201W_allowed = ['T203.W', 'T205.W', 'T224.W', 'T557.W']
G2AR_T202W_allowed = ['T210.R']
G2AR_T203W_allowed = ['T204.R']
G2AR_T204R_allowed = ['T203.W', 'T205.W', 'T224.W', 'T557.W']
G2AR_T205W_allowed = ['T206.W', 'T224.W', 'T557.W']
G2AR_T206W_allowed = ['T203.W', 'T205.W']
G2AR_T210R_allowed = ['T201.W', 'T202.W']
G2AR_T224W_allowed = ['T203.W', 'T205.W', 'T557.W']
G2AR_T555R_allowed = ['T201.W', 'T202.W']
G2AR_T557W_allowed = ['T208.R', 'T210.R']

G2AW_TRANSACTIONS = ['T555.W', 'T201.W', 'T202.W', 'T217.W', 'T218.R', 'T205.W', 'T206.W','T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.W', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T224.W', 'T557.W']
G2AW_T201W_allowed = ['T217.W', 'T205.W', 'T224.W', 'T557.W']
G2AW_T202W_allowed = ['T210.R']
G2AW_T203W_allowed = ['T218.R']
G2AW_T204R_allowed = ['T217.W', 'T205.W', 'T224.W', 'T557.W']
G2AW_T205W_allowed = ['T206.W', 'T224.W', 'T557.W']
G2AW_T206W_allowed = ['T217.W', 'T205.W']
G2AW_T210R_allowed = ['T201.W', 'T202.W']
G2AW_T224W_allowed = ['T217.W', 'T205.W', 'T557.W']
G2AW_T555R_allowed = ['T201.W']
G2AW_T557W_allowed = ['T208.R', 'T210.R']

G2BR_TRANSACTIONS = ['T556.R', 'T201.W', 'T202.W', 'T203.W', 'T204.R', 'T205.W', 'T206.W','T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T224.W', 'T557.W']
G2BR_T201W_allowed = ['T203.W', 'T205.W', 'T224.W', 'T557.W']
G2BR_T202W_allowed = ['T210.R']
G2BR_T203W_allowed = ['T204.R']
G2BR_T204R_allowed = ['T203.W', 'T205.W', 'T224.W', 'T557.W']
G2BR_T205W_allowed = ['T206.W', 'T224.W', 'T557.W']
G2BR_T206W_allowed = ['T203.W', 'T205.W']
G2BR_T210R_allowed = ['T201.W', 'T202.W']
G2BR_T224W_allowed = ['T203.W', 'T205.W', 'T557.W']
G2BR_T555R_allowed = ['T201.W', 'T202.W']
G2BR_T557W_allowed = ['T208.R', 'T210.R']

G2BW_TRANSACTIONS = ['T556.W', 'T201.W', 'T202.W', 'T217.W', 'T218.R', 'T205.W', 'T206.W','T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.W', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T224.W', 'T557.W']
G2BW_T201W_allowed = ['T217.W', 'T205.W', 'T224.W', 'T557.W']
G2BW_T202W_allowed = ['T210.R']
G2BW_T203W_allowed = ['T218.R']
G2BW_T204R_allowed = ['T217.W', 'T205.W', 'T224.W', 'T557.W']
G2BW_T205W_allowed = ['T206.W', 'T224.W', 'T557.W']
G2BW_T206W_allowed = ['T217.W', 'T205.W']
G2BW_T210R_allowed = ['T201.W', 'T202.W']
G2BW_T224W_allowed = ['T217.W', 'T205.W', 'T557.W']
G2BW_T555R_allowed = ['T201.W']
G2BW_T557W_allowed = ['T208.R', 'T210.R']

fake = Faker()
fake.add_provider(Enzyme)

#pick random SPID, METER_MNF_ METER_SERIAL
def pick_spid_meter():
    spid_meter = random.choice(list(SPIDS_METERS.items()))
    spid = spid_meter[0]
    meter_mnf = spid_meter[1][0]
    meter_ser = spid_meter[1][1]
    return spid, meter_mnf, meter_ser

def random_email():
    return fake.company_email()

def random_string():
    return ''.join(random.choice(string.ascii_letters) for _ in range(15))

def random_name():
    return names.get_full_name()

def random_phone():
    return random.randint(4400000000, 4499999999)

def random_meter_ser():
    return '10W' + str(random.randint(0000000000, 9999999999))

def random_meter_mnf():
    return ''.join(random.choice(string.ascii_letters).upper() for _ in range(random.randint(4, 10)))

def random_gisx():
    return random.randint(82644, 655612)

def random_gisy():
    return random.randint(5186, 657421)

def random_meter_loc():
    return random.choice(['UNDER_THE_TREE', 'SOMEWHERE', 'IN_THE_BASEMENT', 'ON_THE_ROOF', 'UNDER_THE_SINK', 'NO_IDEA_WHERE', 'IN_THE_BACKYARD', 'BELOW_THE_WINDOW'])

def date_not_weekend():
    if datetime.today().weekday() >=0 and datetime.today().weekday() <=3:
        return '[today+' + str(4 - datetime.today().weekday()) + ']'
    else:
        return '[today+3]'

def time_not_weekend():
    if datetime.today().weekday() >=0 and datetime.today().weekday() <=3:
        return '[now+' + str(4 - datetime.today().weekday()) + ']'
    else: 
        return '[now+3]'
    
def get_random_address(): # !!!!!!!!!! LEARN HOW TO USE IT
    rand_address =  real_random_address()
    return rand_address["address1"]

#C1
D8036 = ['ERROR', 'DUPLICATE', 'SWITCHED', 'REJECTION', 'UNABLEASST', 'DISAGREEPLAN'] # T211.R T211.W Cancellation Reason Code
D8226 = ['NOCONTACT', 'UNCOOPCUST', 'INACCONTACT', 'MOREDETAILS'] # T203.W T217W Additional Information Request Code
D8228 = ['WHOL', 'NONWHOL'] # T206.W Site Visit Failure Code
D8229 = ['CUSTOMER', 'RETAILER', 'THIRDPARTY', 'CONSENTS', 'REGULAT', 'WEATHER', 'FORCEMAJ', 'INFOREQD', 'BULK'] # T213.W Request Deferral Code
D8230 = ['INACCURATE', 'DUPLICATE', 'WRONGPRO', 'POLICY', 'HOUSEHOLD', 'NOTWHOL'] # T202.W Reject Reason Code
D8231 = ['DISPREJECT', 'DISPCMOS'] # T210.R Resubmit Reason Code
D8236 = ['EMAIL', 'TEL', 'BOTH'] # T321.R T321.W Customer Preferred Method of Contact
D2005 = ['SEMDV', 'NA'] # T321.R T321.W Customer Classification – Sensitive Customer
D8237 = ['AM', 'PM', 'BOTH'] # T321.R T321.W
D8242 = ['METER', 'SUPPLY', 'BOTH'] # T321.R T321.W
D8262 = ['ACCEPT', 'REJECT'] # T321.R T321.W
D8242 = ['PDF', 'JPG', 'PNG'] # T215.R T215.W
D3025 = ['I', 'O']

#B1 #B7
D8327 = ['NEWINSTALL', 'CHGNEW', 'LOCCHGNEW', 'LOCCHGEXG', 'UNFEASIBLE'] # B1 B7 COMPLETED T223.W -> Meter Work Complete Code

#B3
D8346 = ['INSIDE', 'OUTSIDE', 'UNKNOWN'] # T353.R Meter Location Code
D8348 = ['OVERRECORD', 'UNDERRECORD', 'OTHER'] # T353.R Meter Location Code
D8367 = ['AFTEREXCHG', 'ALREADYTESTSED', 'INSITUTESTED']
D8368 = ['WITHIN', 'OUTSIDE']
D8369 = ['1', '0']

#B5
D8227 = ['PARTS', 'STREETWORKS', 'THIRDPARTY', 'CUSTCONFRM', 'PREPWORK', 'OTHER'] # T224.W Delay Reason Code - B5R B5W - Advise Process Delay
D8332 = ['NOISSUE', 'NOWATER', 'FLOODING'] # T351.R T351.W Public Health Issue
D8333 = ['REMOVED', 'NOTREMOVED'] # T351.R T351.W Datalogger Status
D8335 = ['STD', 'NONSTD'] # T351.R T351.W Meter Model
D8337 = D8838 = D8839 = ['STOPPED', 'BACKWARD', 'SLOWED', 'BURRIED', 'CONDENS', 'ELECT', 'BURST', 'SMASHED', 'REMOVED', 'NONMETER', 'OTHER'] # T351.R T351.W Meter Fault
D8330 = ['0', '1'] #T351.R T351.W Meter Fault address same as CMOS
D8341 = ['REPLACED', 'REPAIRED', 'NOUPDATE', 'NOFAULT', 'NONMETER', 'UPDATE'] # T352.W Complete Reason Code

#B7
D8326 = ['CHGTYPE', 'CHGSSIZE', 'CHGLSIZE', 'CHGLOC'] # B7 T365.R Request Meter Change -> Meter Work Request Type

#F4
D8364 = ['DWENQUIRY', 'OTHERENQUIRY'] # F4 T505.R Request Type
D8365 = ['WATERQUALITY', 'FLUORIDE', 'HARDNESS', 'QUALITYREPT', 'GENERAL', 'ANIMALS', 'LEAD', 'PUBLICINFO'] # F4 T505.R Drinking Water Enquiry Type - D8364 = 'DWENQUIRY'
D8352 = ['FOLLOWON', 'NOFOLLOWON'] #F4 T222.W Response Type

#F5
D8356 = ['FIRST', 'FURTHER', 'CCWLEVEL', 'ADR', 'OTHER'] #F5 T501.R/W Complaint Level
D8358 = ['ADMINISTRATION', 'METERINGASSET', 'BILLING', 'WATER', 'SEWERAGE', 'OTHER'] #F5 T501.R/W Complaint Category
D8360 = ['GSSFAILURE', 'OTHER', 'NONE'] #F5 T501.R/W Compensation Claimed

#G2A G2B
D8371 = ['NEWCONSENT', 'NEWTEMPCONSENT', 'RENEWCONSENT']
D8374 =	['YES', 'NO', 'NA']
D8375 =	['YES', 'NO', 'NA']
D8376 =	['YES', 'NO', 'NA']
D8377 =	['YES', 'NO', 'NA']
D8378 =	['YES', 'NO', 'NA']
D8379 =	['YES', 'NO', 'NA']
D8380 =	['YES', 'NO', 'NA']
D8381 =	['YES', 'NO', 'NA']
D8382 =	['NOTREQD', 'GRANTED', 'NOTGRANTED']
D8383 =	['PERMANENT', 'TEMPORARY', 'RENEWAL']

test_case_sequence = []
def generate_test_case(loop_times):
    global test_case_sequence
    new_filename = ""
    program_mode = input("Do you want to run in [I]interactive or [P]redefined mode? ")
    while program_mode not in ('I', 'P'):
        program_mode = input("You can choose only from [I]interactive or [P]redefined mode? What's your choice? ")
    if program_mode == 'I':
        ############################################################### ASK USER WHICH TEST CASE HE/SHE WANTS
        available_processes = ""
        for b in range(len(PROC_NAMES)):
            ################################################################### get process name from PROC_NAMES  + get process description
            available_processes = available_processes + "{:2}".format(b+1) + " - " + list(PROC_NAMES.keys())[b] + " - " +  PROC_NAMES.get(list(PROC_NAMES.keys())[b])+ "\n"
        chosen_process = input("Choose process to start with - available are: \n\n" + available_processes + "\n")
        while ((not chosen_process.isdigit()) or (int(chosen_process) not in range(1, len(PROC_NAMES)+1))):
                chosen_process = input("\nWrong choice! You can only choose from available processes: \n" + available_processes + "\nChoose process:")
        chosen_proc = int(chosen_process)
        print("\nYour choice: " + str(list(PROC_NAMES.keys())[chosen_proc-1]) + " - " +  PROC_NAMES.get(PROCESSES[chosen_proc-1])) 
        
        #choose 1st transaction from chosen process i.e. C1R + _TRANSACTIONS
        chosen_process1 = globals()[PROCESSES[chosen_proc-1] + '_TRANSACTIONS']
        starting_transaction = chosen_process1[0]
        
        #append initiating transaction to the TEST_CASE_SEQUENCE
        test_case_sequence.append(starting_transaction)
        more_transactions = input("I will generate test case with transaction:\n" + starting_transaction + " - " + TRANSACTION_NAMES.get(starting_transaction) + "\nDo you want to add more transactions for current process? [Y]/[N], [C]hange process or [A]ny transaction from current process?\n")
        while more_transactions not in ('Y', 'N', 'C', 'A'):
            more_transactions = input("You can only choose [Y]es, [N]o, [C]hange or [A]ny. Do you want to add more transactions for current process? [Y]/[N], [C]hange process or [A]ny transaction from current process?\n")
        if more_transactions =='N':
            print("Generating a test case, thank you. Bye!")
        if more_transactions == 'C':
            generate_test_case(max_loop) # REPEAT FROM START - TEST_CASE_SEQUENCE WILL NOT BE OVERWRITTEN!!!!    
            
        # when A selected loop through all transactions in the process, when Y selected loop through available transactions in the process    
        while more_transactions == 'Y' or more_transactions == 'A':
            if  more_transactions == 'A':
                next_transactions = globals()[PROCESSES[chosen_proc-1] + '_TRANSACTIONS']
            if more_transactions == 'Y':   
                next_transactions = globals()[PROCESSES[chosen_proc-1] + '_' + starting_transaction.replace('.', '') + '_allowed']
            print("\nAvailable transactions for this process are:")
            next_transactions1 = ""
            for i in range(len(next_transactions)):
                next_transactions1 += str(i+1) + " " + next_transactions[i] + "\n" 
                print(str(i+1) + " " + str(next_transactions[i]) + " - " + TRANSACTION_NAMES.get(next_transactions[i]))
            print('\n')
            next_transaction = input("Which transaction you want next?\n")
            #allow user to choose only valid transactions - check if user inpur is digit and it is from allowed range
            while ((not next_transaction.isdigit()) or (int(next_transaction) not in range(1, len(next_transactions)+1))):
                    next_transaction = input("You can only choose from available transactions \n" + next_transactions1  + "\nChoose transaction: ")
            next_tran = int(next_transaction)
            print("Your choice: " + str(next_transactions[next_tran-1])) 
            test_case_sequence.append(next_transactions[next_tran-1])
            # if in available processes mode exit when found T208.R
            if more_transactions == 'Y' and str(next_transactions[next_tran-1]) == 'T208.R':
                print("\nTest case sequence:")
                print(test_case_sequence)
                print("\nT208.R was the last transaction. Generating a test case, thank you. Bye!")
                break
            print("\nTest case sequence:")
            print(test_case_sequence)
            starting_transaction = next_transactions[next_tran-1]
            more_transactions = input("Do you want to add more transactions for current process? [Y]/[N], [C]hange process or [A]ny transaction from current process?\n")
            while more_transactions not in ('Y', 'N', 'C', 'A'):
                more_transactions = input("You can only choose [Y]es, [N]o, [C]hange or [A]ny. Do you want to add more transactions for current process? [Y]/[N], [C]hange process or [A]ny transaction from current process?\n")
            if more_transactions == 'C':
                generate_test_case(max_loop) # REPEAT FROM START - TEST_CASE_SEQUENCE WILL NOT BE OVERWRITTEN!!!!                   
            if more_transactions =='N':
                print("Generating a test case, thank you. Bye!") 
    if program_mode == 'P':
        test_case_sequence = TEST_CASE_TRANSACTIONS        
##############################################################################    
    for a in range(loop_times):
        # assign random SPID, METER_MNF, METER_SERIAL to variables - use EXCEL or SPIDS_METERS static dictionary
        SPID, METER_MNF, METER_SER = pick_spid_meter()
        CUST_EMAIL = random_email()
        RET_EMAIL = random_email()
        RANDOM_STRING = random_string()
        CUST_RANDOM_NAME = random_name()
        CUST_RANDOM_PHONE = random_phone()
        RET_RANDOM_NAME = random_name()
        RET_RANDOM_PHONE = random_phone()
        CUST_RANDOM_NAME2 = random_name()
        CUST_RANDOM_PHONE2 = random_phone()
        RET_RANDOM_NAME2 = random_name()
        RET_RANDOM_PHONE2 = random_phone()
        RANDOM_METER_SER = random_meter_ser()
        RANDOM_METER_MNF = random_meter_mnf()
        RANDOM_GISX = random_gisx()
        RANDOM_GISY = random_gisy()
        OUTR_RANDOM_GISX = random_gisx()
        OUTR_RANDOM_GISY = random_gisy()
        RANDOM_METER_LOC = random_meter_loc()
        RANDOM_OUTRE_LOC = random_meter_loc()
        DATE_NOT_WEEKEND = date_not_weekend()
        TIME_NOT_WEEKEND = time_not_weekend()
        RANDOM_ADDRESS1 = get_random_address()
        RANDOM_ADDRESS2 = get_random_address()
        RANDOM_ADDRESS3 = get_random_address()
        RANDOM_ADDRESS4 = get_random_address()
        RANDOM_ADDRESS5 = get_random_address()

        T321R_data_items = [# basic data
                        SPID, 'MEASURED', 'RET_' + RANDOM_STRING, '', '1', '[today]', '1',
                        # customer and retailer data
                        CUST_RANDOM_NAME, CUST_RANDOM_PHONE, '105', CUST_RANDOM_NAME2, CUST_RANDOM_PHONE2, '122', CUST_EMAIL, '1', 'EMAIL', random.choice(D8237),
                        RANDOM_STRING, random.choice(D2005), RANDOM_STRING, RET_RANDOM_NAME, RET_RANDOM_PHONE, '210', RET_RANDOM_NAME2, RET_RANDOM_PHONE2, '224', RET_EMAIL,
                        # first meter details
                        METER_MNF, METER_SER, '0', '120', '[today-1]', '0', 'METER', '1', RANDOM_METER_MNF, '1', RANDOM_METER_SER, '1', '5', '1', '5',
                        '1', RANDOM_GISX, '1', RANDOM_GISY, '1', 'O', '1', RANDOM_METER_LOC, '1', OUTR_RANDOM_GISX, '1', OUTR_RANDOM_GISY, '1', 'I', '1', RANDOM_OUTRE_LOC, RANDOM_STRING, '',
                        # second meter details empty
                        '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                        # missing meters
                        '', '', '', '',
                        # unmeasrued data empty
                        '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''
                        ]
        T321W_data_items = [# basic data
                        SPID, 'MEASURED', '', '[today]',
                        # first meter details
                        METER_MNF, METER_SER, '0', '120', '[today-1]', '0', 'METER', '1', RANDOM_METER_MNF, '1', RANDOM_METER_SER, '1', '5', '1', '5',
                        '1', RANDOM_GISX, '1', RANDOM_GISY, '1', 'O', '1', RANDOM_METER_LOC, '1', OUTR_RANDOM_GISX, '1', OUTR_RANDOM_GISY, '1', 'I', '1', RANDOM_OUTRE_LOC, RANDOM_STRING, '',
                        # second meter details empty
                        '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                        # missing meters
                        '', '', '', '',
                        # unmeasrued data empty
                        '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''
                        ]
        T201W_data_items = ['[orid]', 'ACCEPTED']
        T202W_data_items = ['[orid]', 'WSL-123456', random.choice(D8230), 'REJECTED']
        T203W_data_items = ['[orid]', random.choice(D8226), 'INFOREQST']
        T204R_data_items = ['[orid]', 'INFOPROVD']
        T205W_data_items = ['[orid]', DATE_NOT_WEEKEND, '', 'VISITSCHED']
        T206W_data_items = ['[orid]', random.choice(D8228), 'VISITNOTCOMP']
        T207R_data_items = ['[orid]', 'RETAILER_COMMENT']
        T207W_data_items = ['[orid]', 'WHOLESALER_COMMENT']
        T208R_data_items = ['[orid]', 'CLOSED']
        T210R_data_items = ['[orid]', random.choice(D8231), 'RESUBMITTED']
        T211R_data_items = ['[orid]', random.choice(D8036), 'RTL CANCELLED']
        T211W_data_items = ['[orid]', random.choice(D8036), 'WSL CANCELLED']
        T212W_data_items = ['[orid]', 'PREPPLAN']
        T213W_data_items = ['[orid]', random.choice(D8229), '[today]', '[today+7]', 'START_DEFERRAL']
        T214W_data_items = ['[orid]', '[today]', 'END_DEFERRAL'] # can think of function to peek working day
        T215R_data_items = ['[orid]', '', 'img1png', 'PNG', '4oCwUE5HChoKICAgCklIRFIgICADICAgAwgCICAgxa5KIsSNICAgCXBIWXMgIA7DhCAgDsOEAeKAoisOGyAgICdJREFUCOKEomPDlG7Dn8O2y5nLmX8BNnbCpn/LmcWjMTMzH8OrxI9nYmXLmXd5w7cmxLrFmBAgxZDFnwrFpH4uJsKsICAgIElFTkTCrkJg4oCa']
        T215W_data_items = ['[orid]', '', 'img1png', 'PNG', '4oCwUE5HChoKICAgCklIRFIgICADICAgAwgCICAgxa5KIsSNICAgCXBIWXMgIA7DhCAgDsOEAeKAoisOGyAgICdJREFUCOKEomPDlG7Dn8O2y5nLmX8BNnbCpn/LmcWjMTMzH8OrxI9nYmXLmXd5w7cmxLrFmBAgxZDFnwrFpH4uJsKsICAgIElFTkTCrkJg4oCa']
        T216R_data_items = ['[orid]', T216_URL]
        T216W_data_items = ['[orid]', T216_URL]
        T217W_data_items = ['[orid]', '1', random.choice(D8226), 'CUSTINFOREQST']

        T218R_data_items = ['[orid]', 'RET_' + RANDOM_STRING, '[today]',
                             '1', CUST_RANDOM_NAME, CUST_RANDOM_PHONE, '105', CUST_RANDOM_NAME2, CUST_RANDOM_PHONE2, '122', CUST_EMAIL, '1', 'EMAIL', 
                             random.choice(D8237),  RANDOM_STRING, random.choice(D2005), RANDOM_STRING,RET_RANDOM_NAME, RET_RANDOM_PHONE, '210', 
                             RET_RANDOM_NAME2, RET_RANDOM_PHONE2, '224', RET_EMAIL, 'CUSTINFOPROVD']
        T220W_data_items = ['[orid]', fake.paragraph(nb_sentences=1)]
        T221R_data_items = ['[orid]', fake.paragraph(nb_sentences=1)]
        T222W_data_items = [# basic data
                            '[orid]', 'FOLLOWON', fake.paragraph(nb_sentences=1), fake.paragraph(nb_sentences=1),'[today+' + str(random.randint(2, 10))  +']', '1', '', '', fake.paragraph(nb_sentences=1)
                            ]
        T223W_data_items = [# basic data
                            '[orid]', 'NEWINSTALL', fake.paragraph(nb_sentences=1),
                            #meter work copletion data
                            RANDOM_METER_MNF, RANDOM_METER_SER, random.randint(100, 1000), '[today-' + str(random.randint(2, 30))  +']', random.randint(5, 15), random.randint(2, 6), RANDOM_GISX, RANDOM_GISY, random.choice(D3025), RANDOM_METER_LOC, OUTR_RANDOM_GISX, OUTR_RANDOM_GISY, random.choice(D3025), RANDOM_OUTRE_LOC, RANDOM_STRING, RANDOM_STRING, RANDOM_ADDRESS1, RANDOM_ADDRESS2, RANDOM_ADDRESS3, RANDOM_ADDRESS4, RANDOM_ADDRESS5, random.choice(POSTCODES), random.randint(1, 99999999)
                            ] 
        T224W_data_items = ['[orid]', random.choice(D8227), 'PROCDELAY_' + RANDOM_STRING]
        T323W_data_items = ['[orid]', 'ABLE', TIME_NOT_WEEKEND, '0', '1', DATE_NOT_WEEKEND, 'PLANPROP']
        T324R_data_items = ['[orid]', 'PLANAGREED']
        T325R_data_items = ['[orid]', 'PLANDISP']
        T322W_data_items = [# basic data
                            '[orid]', '1', 
                            # 1st meter data
                            METER_MNF, METER_SER, '0', '1', RANDOM_METER_MNF, '1', RANDOM_METER_SER, '1', '12', '1', '5',
                            '1', RANDOM_GISX, '1', RANDOM_GISY, '1', 'I', '1', RANDOM_METER_LOC, '1', OUTR_RANDOM_GISX, '1', OUTR_RANDOM_GISY, '1', 'O', '1', RANDOM_OUTRE_LOC, 'MORE_INFO_T322W',
                            # 2nd meter data
                            '', '', '', '', '', '', '', '', '', '', '',
                            '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                            # missing meters data
                            '', '', '', '', '', '',
                            # unmeasured  data
                            '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 
                            '', '', '']
        T351R_data_items = [# basic data
                            SPID, 'RET_' + RANDOM_STRING, '', '1', '[today]', '1',
                            # customer and retailer data
                            CUST_RANDOM_NAME, CUST_RANDOM_PHONE, '105', CUST_RANDOM_NAME2, CUST_RANDOM_PHONE2, '122', CUST_EMAIL, '1', 'EMAIL', random.choice(D8237),
                            RANDOM_STRING, random.choice(D2005), RANDOM_STRING, RET_RANDOM_NAME, RET_RANDOM_PHONE, '210', RET_RANDOM_NAME2, RET_RANDOM_PHONE2, '224', RET_EMAIL,
                            # meter details
                            METER_MNF, METER_SER, random.choice(D8330), str(random.randint(1111, 9999)) + RANDOM_STRING, str(random.randint(1111, 9999)) + RANDOM_STRING, RANDOM_ADDRESS1, RANDOM_ADDRESS2, RANDOM_ADDRESS3, RANDOM_ADDRESS4, RANDOM_ADDRESS5, random.choice(POSTCODES), random.randint(1, 99999999), '0', '',
                            random.choice(D8332), '1', 'NOTREMOVED', '[today+' + str(random.randint(1, 15))  +']', random.choice(D8335), 'METER_' + RANDOM_STRING, 'STOPPED', 'BACKWARD', 'SLOWED', '', random.randint(100, 9999), '[today-' + str(random.randint(1, 45))  +']', fake.paragraph(nb_sentences=1)
                            ]
        T351W_data_items = [# basic data
                            SPID, '', '[today]',
                            # meter details
                            METER_MNF, METER_SER, random.choice(D8330), str(random.randint(1111, 9999)) + RANDOM_STRING, str(random.randint(1111, 9999)) + RANDOM_STRING, RANDOM_ADDRESS1, RANDOM_ADDRESS2, RANDOM_ADDRESS3, RANDOM_ADDRESS4, RANDOM_ADDRESS5, random.choice(POSTCODES), random.randint(1, 99999999), '0', '',
                            random.choice(D8332), '1', 'NOTREMOVED', '[today+' + str(random.randint(1, 15))  +']', random.choice(D8335), 'METER_' + RANDOM_STRING, 'STOPPED', 'BACKWARD', 'SLOWED', '', random.randint(100, 9999), '[today-' + str(random.randint(1, 45))  +']', fake.paragraph(nb_sentences=1)
                            ]
        T352W_data_items = [# basic data
                            '[orid]', 'REPLACED', # random.choice(D8341), # when other option than REPLACED no NEW METER can be provided - CONSIDER RULE!
                            # meter data
                            METER_MNF, METER_SER, RANDOM_METER_MNF, RANDOM_METER_SER, '120', '[today]', '1', '12', '1', '5', '1', RANDOM_GISX, '1', RANDOM_GISY, '1', 'I', '1', RANDOM_METER_LOC,
                            '1', OUTR_RANDOM_GISX, '1', OUTR_RANDOM_GISY, '1', 'O', '1', RANDOM_OUTRE_LOC, 'UPDATED_METER_T352W',
                            ]
        T353R_data_items = [# basic data
                            SPID, 'RET_' + RANDOM_STRING, '', '1', '[today]', 
                            # customer and retailer data
                            '1', CUST_RANDOM_NAME, CUST_RANDOM_PHONE, '105', CUST_RANDOM_NAME2, CUST_RANDOM_PHONE2, '122', CUST_EMAIL, '1', 'EMAIL', random.choice(D8237),
                            RANDOM_STRING, random.choice(D2005), RANDOM_STRING, RET_RANDOM_NAME, RET_RANDOM_PHONE, '210', RET_RANDOM_NAME2, RET_RANDOM_PHONE2, '224', RET_EMAIL,
                            # install meter details
                            '1', RANDOM_STRING, RANDOM_STRING,  RANDOM_ADDRESS1, RANDOM_ADDRESS2, RANDOM_ADDRESS3, RANDOM_ADDRESS4, RANDOM_ADDRESS5, random.choice(POSTCODES), random.randint(1, 99999999), random.randint(10, 30), random.choice(D8335), 'METER_' + RANDOM_STRING, random.choice(D8346), fake.paragraph(nb_sentences=1), fake.paragraph(nb_sentences=1)
                            ]
        T355R_data_items = [# basic data
                            SPID, 'RET_' + RANDOM_STRING, '', '1', '[today]', 
                            # customer and retailer data
                            '1', CUST_RANDOM_NAME, CUST_RANDOM_PHONE, '105', CUST_RANDOM_NAME2, CUST_RANDOM_PHONE2, '122', CUST_EMAIL, '1', 'EMAIL', random.choice(D8237),
                            RANDOM_STRING, random.choice(D2005), RANDOM_STRING, RET_RANDOM_NAME, RET_RANDOM_PHONE, '210', RET_RANDOM_NAME2, RET_RANDOM_PHONE2, '224', RET_EMAIL,
                            # request accuracy meter details
                            METER_MNF, METER_SER, '1', RANDOM_STRING, RANDOM_STRING, RANDOM_ADDRESS1, RANDOM_ADDRESS2, RANDOM_ADDRESS3, RANDOM_ADDRESS4, RANDOM_ADDRESS5, random.choice(POSTCODES), random.randint(1, 99999999), '0', '', '1', 'NOTREMOVED', '[today+' + str(random.randint(1, 15))  +']', random.choice(D8335), 'METER_' + RANDOM_STRING, random.choice(D8348), fake.paragraph(nb_sentences=1), fake.paragraph(nb_sentences=1) 
                            ]
        T355W_data_items = [# basic data
                            SPID, '', '[today]', 
                            # request accuracy meter details
                            METER_MNF, METER_SER, '1', RANDOM_STRING, RANDOM_STRING, RANDOM_ADDRESS1, RANDOM_ADDRESS2, RANDOM_ADDRESS3, RANDOM_ADDRESS4, RANDOM_ADDRESS5, random.choice(POSTCODES), random.randint(1, 99999999), '0', '', '1', 'NOTREMOVED', '[today+' + str(random.randint(1, 15))  +']', random.choice(D8335), 'METER_' + RANDOM_STRING, random.choice(D8348), fake.paragraph(nb_sentences=1), fake.paragraph(nb_sentences=1)
                            ]
        T356W_data_items = [# basic data
                            '[orid]',
                            #rAwaiting Meter Accuracy Test data
                            METER_MNF, METER_SER, random.choice(D8367), random.choice(D8368), '1', random.randint(100, 9999), random.randint(100, 9999), '[today-' + str(random.randint(1, 45))  +']', '1', '', fake.paragraph(nb_sentences=1)
                            ]
        T357W_data_items = [# basic data
                            '[orid]',
                            # rAwaiting Meter Accuracy Test data
                            METER_MNF, METER_SER, RANDOM_METER_MNF, RANDOM_METER_SER, random.randint(100, 9999), '[today-' + str(random.randint(1, 45))  +']', '1', random.randint(3, 10), '1', random.randint(2, 6), '1', RANDOM_GISX, '1', RANDOM_GISY, '1', random.choice(D3025), '1', RANDOM_METER_LOC, '1', OUTR_RANDOM_GISX, '1', OUTR_RANDOM_GISY, '1', random.choice(D3025), '1', RANDOM_OUTRE_LOC, fake.paragraph(nb_sentences=1)
                            ]
        T365R_data_items = [# basic data
                            SPID, 'RET_' + RANDOM_STRING, '', '1', '[today]', 
                            # customer and retailer data
                            '1', CUST_RANDOM_NAME, CUST_RANDOM_PHONE, '105', CUST_RANDOM_NAME2, CUST_RANDOM_PHONE2, '122', CUST_EMAIL, '1', 'EMAIL', random.choice(D8237),
                            RANDOM_STRING, random.choice(D2005), RANDOM_STRING, RET_RANDOM_NAME, RET_RANDOM_PHONE, '210', RET_RANDOM_NAME2, RET_RANDOM_PHONE2, '224', RET_EMAIL,
                            # Meter Change details
                            METER_MNF, METER_SER, '1', RANDOM_STRING, RANDOM_STRING,  RANDOM_ADDRESS1, RANDOM_ADDRESS2, RANDOM_ADDRESS3, RANDOM_ADDRESS4, RANDOM_ADDRESS5, random.choice(POSTCODES), random.randint(1, 99999999), '0', '', random.choice(D8332), '1', 'NOTREMOVED', '[today+' + str(random.randint(1, 15))  +']', random.choice(D8326), random.randint(10, 30), random.choice(D8335), 'METER_' + RANDOM_STRING, random.choice(D8346), fake.paragraph(nb_sentences=1), fake.paragraph(nb_sentences=1)
                            ]
        T501R_data_items = [# basic data
                            SPID, 'RET_' + RANDOM_STRING, '[today-' + str(random.randint(0, 7))  +']', fake.paragraph(nb_sentences=1), random.choice(D8356), RANDOM_STRING, random.choice(D8358), RANDOM_STRING,random.choice(D8360), RANDOM_STRING, '', '1', fake.paragraph(nb_sentences=1), '[today]', '1',
                            # customer and retailer data
                            CUST_RANDOM_NAME, CUST_RANDOM_PHONE, '105', CUST_RANDOM_NAME2, CUST_RANDOM_PHONE2, '122', CUST_EMAIL, '1', 'EMAIL', random.choice(D8237),
                            fake.paragraph(nb_sentences=1), random.choice(D2005), fake.paragraph(nb_sentences=1), RET_RANDOM_NAME, RET_RANDOM_PHONE, '210', RET_RANDOM_NAME2, RET_RANDOM_PHONE2, '224', RET_EMAIL,
                            ]
        T501W_data_items = [# basic data
                            SPID, '[today-' + str(random.randint(0, 7))  +']', fake.paragraph(nb_sentences=1), random.choice(D8356), ''.join(random.choice(string.ascii_letters) for _ in range(15)), random.choice(D8358),''.join(random.choice(string.ascii_letters) for _ in range(15)),random.choice(D8360), ''.join(random.choice(string.ascii_letters) for _ in range(15)), '', fake.paragraph(nb_sentences=1), '[today]'      # [today - 0] = [today]!!!
                           ]
        T505R_data_items = [# basic data
                            SPID, 'RET_' + RANDOM_STRING, '[today-' + str(random.randint(0, 7))  +']', 'DWENQUIRY', random.choice(D8365), fake.enzyme(), '', '1', fake.paragraph(nb_sentences=1), '[today]', '1',       # [today - 0] = [today]!!!
                            # customer and retailer data
                            CUST_RANDOM_NAME, CUST_RANDOM_PHONE, '105', CUST_RANDOM_NAME2, CUST_RANDOM_PHONE2, '122', CUST_EMAIL, '1', 'EMAIL', random.choice(D8237),
                            fake.paragraph(nb_sentences=1), random.choice(D2005), fake.paragraph(nb_sentences=1), RET_RANDOM_NAME, RET_RANDOM_PHONE, '210', RET_RANDOM_NAME2, RET_RANDOM_PHONE2, '224', RET_EMAIL,
                            ]
        T505W_data_items = [# basic data
                            SPID, '[today-' + str(random.randint(0, 7))  +']', 'DWENQUIRY',random.choice(D8365),  fake.enzyme(), '', fake.paragraph(nb_sentences=1),   '[today]', # [today - 0] = [today]!!!
                            ]
        T551R_data_items = [# basic data
                            SPID, 'DPID_' + RANDOM_STRING, 'RET_' + RANDOM_STRING, '[today-' + str(random.randint(0, 7))  +']', fake.paragraph(nb_sentences=1), '', '1',  fake.paragraph(nb_sentences=1), '[today]', '1',    # [today - 0] = [today]!!!
                            # customer and retailer data
                            CUST_RANDOM_NAME, CUST_RANDOM_PHONE, '105', CUST_RANDOM_NAME2, CUST_RANDOM_PHONE2, '122', CUST_EMAIL, '1', 'EMAIL', random.choice(D8237),
                            fake.paragraph(nb_sentences=1), random.choice(D2005), fake.paragraph(nb_sentences=1), RET_RANDOM_NAME, RET_RANDOM_PHONE, '210', RET_RANDOM_NAME2, RET_RANDOM_PHONE2, '224', RET_EMAIL,
                            ]
        T551W_data_items = [# basic data
                            SPID, 'DPID_' + RANDOM_STRING, '[today-' + str(random.randint(0, 7))  +']', fake.paragraph(nb_sentences=1), '', fake.paragraph(nb_sentences=1), '[today]'      # [today - 0] = [today]!!!
                           ]
        T555R_data_items = [# basic data
                            SPID, 'DPID_' + RANDOM_STRING, 'RET_' + RANDOM_STRING, '', '1', random.choice(D8371), '[today+' + str(random.randint(0, 7))  +']', '[today-' + str(random.randint(0, 7))  +']', fake.paragraph(nb_sentences=1),'[today]',
                            # customer and retailer data
                            '1', CUST_RANDOM_NAME, CUST_RANDOM_PHONE, '105', CUST_RANDOM_NAME2, CUST_RANDOM_PHONE2, '122', CUST_EMAIL, '1', 'EMAIL', random.choice(D8237),
                            fake.paragraph(nb_sentences=1), random.choice(D2005), fake.paragraph(nb_sentences=1), RET_RANDOM_NAME, RET_RANDOM_PHONE, '210', RET_RANDOM_NAME2, RET_RANDOM_PHONE2, '224', RET_EMAIL,
                            # Group Attached Application
                            random.choice(D8374), random.choice(D8375), random.choice(D8376), random.choice(D8377), random.choice(D8378), random.choice(D8379), random.choice(D8380), 'Yes' #random.choice(D8381), - at least 1 Yes
                            ]
        T555W_data_items = [# basic data
                            SPID, 'DPID_' + RANDOM_STRING, 'RET_' + RANDOM_STRING, '', random.choice(D8371), '[today+' + str(random.randint(0, 7))  +']', '[today-' + str(random.randint(0, 7))  +']', fake.paragraph(nb_sentences=1),'[today]',
                            # Group Attached Application
                            random.choice(D8374), random.choice(D8375), random.choice(D8376), random.choice(D8377), random.choice(D8378), random.choice(D8379), random.choice(D8380), 'Yes' #random.choice(D8381), - at least 1 Yes
                            ]
        T556R_data_items = [# basic data
                            'RET_' + RANDOM_STRING, '1', '', 'MOSLTEST-W', 'MOSLTEST-W', 'MOSLTEST-W', 'SEC_' + RANDOM_STRING, 'PRI_' + RANDOM_STRING, RANDOM_ADDRESS1, RANDOM_ADDRESS2, RANDOM_ADDRESS3, RANDOM_ADDRESS4, RANDOM_ADDRESS5, random.choice(POSTCODES), random.randint(1, 99999999), RANDOM_STRING, random.randint(1, 99999999), random.choice(D8371), '[today-' + str(random.randint(0, 7))  +']', fake.paragraph(nb_sentences=1),'[today]',
                            # customer and retailer data
                            '1', CUST_RANDOM_NAME, CUST_RANDOM_PHONE, '105', CUST_RANDOM_NAME2, CUST_RANDOM_PHONE2, '122', CUST_EMAIL, '1', 'EMAIL', random.choice(D8237),
                            fake.paragraph(nb_sentences=1), random.choice(D2005), fake.paragraph(nb_sentences=1), RET_RANDOM_NAME, RET_RANDOM_PHONE, '210', RET_RANDOM_NAME2, RET_RANDOM_PHONE2, '224', RET_EMAIL,
                            # Group Attached Application
                            random.choice(D8374), random.choice(D8375), random.choice(D8376), random.choice(D8377), random.choice(D8378), random.choice(D8379), random.choice(D8380), 'Yes' #random.choice(D8381), - at least 1 Yes
                            ]
        T556W_data_items = [# basic data
                            'RET_' + RANDOM_STRING, '1', '', 'MOSLTEST-W', 'MOSLTEST-W', 'MOSLTEST-W', 'SEC_' + RANDOM_STRING, 'PRI_' + RANDOM_STRING, RANDOM_ADDRESS1, RANDOM_ADDRESS2, RANDOM_ADDRESS3, RANDOM_ADDRESS4, RANDOM_ADDRESS5, random.choice(POSTCODES), random.randint(1, 99999999), RANDOM_STRING, random.randint(1, 99999999), random.choice(D8371), '[today-' + str(random.randint(0, 7))  +']', fake.paragraph(nb_sentences=1),'[today]',
                            # Group Attached Application
                            random.choice(D8374), random.choice(D8375), random.choice(D8376), random.choice(D8377), random.choice(D8378), random.choice(D8379), random.choice(D8380), 'Yes' #random.choice(D8381), - at least 1 Yes
                            ]
        T557W_data_items = [# basic data
                            '[orid]', 'GRANTED', 'PERMANENT', '[today]', fake.paragraph(nb_sentences=1)
                            ]
        
        TEST_CASE_LENGTH = len(test_case_sequence)
        #gererate test case sequence in Excel file       
        for i in range(TEST_CASE_LENGTH):
            # build file name based on transactions chain. i.e. T321R_T201W_T322W....
            new_filename = new_filename + test_case_sequence[i] + '_'
            # if transaction has .R in the name, it is MOSLTEST-R as requestor
            # put the transaction Source Org ID in sheet Test Case Sequence, column C - Source ID
            ws11.cell(row=i+4+(a*TEST_CASE_LENGTH), column=5).value = test_case_sequence[i]
            if test_case_sequence[i][-1] == 'R':
                ws11.cell(row=i+4+(a*TEST_CASE_LENGTH), column=3).value = RETAILER
            else:
                ws11.cell(row=i+4+(a*TEST_CASE_LENGTH), column=3).value = WHOLESALER
            # then in second sheet 'Test case data' depending on the transaction, insert respctive data items
            match test_case_sequence[i]:
                case 'T321.R':
                    for k in range(len(T321R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T321R_data_items[k]
                case 'T321.W':
                    for k in range(len(T321W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T321W_data_items[k]
                case 'T201.W':
                    for k in range(len(T201W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T201W_data_items[k]
                case 'T202.W':
                    for k in range(len(T202W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T202W_data_items[k]
                case 'T203.W':
                    for k in range(len(T203W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T203W_data_items[k]
                case 'T204.R':
                    for k in range(len(T204R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T204R_data_items[k]
                case 'T205.W':
                    for k in range(len(T205W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T205W_data_items[k]
                case 'T206.W':
                    for k in range(len(T206W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T206W_data_items[k]
                case 'T207.R':
                    for k in range(len(T207R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T207R_data_items[k]
                case 'T207.W':
                    for k in range(len(T207W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T207W_data_items[k]
                case 'T208.R':
                    for k in range(len(T208R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T208R_data_items[k]
                case 'T210.R':
                    for k in range(len(T210R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T210R_data_items[k]
                case 'T211.R':
                    for k in range(len(T211R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T211R_data_items[k]
                case 'T211.W':
                    for k in range(len(T211W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T211W_data_items[k]
                case 'T212.W':
                    for k in range(len(T212W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T212W_data_items[k]
                case 'T213.W':
                    for k in range(len(T213W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T213W_data_items[k]
                case 'T214.W':
                    for k in range(len(T214W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T214W_data_items[k]
                case 'T215.R':
                    for k in range(len(T215R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T215R_data_items[k]
                case 'T215.W':
                    for k in range(len(T215W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T215W_data_items[k]
                case 'T216.R':
                    for k in range(len(T216R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T216R_data_items[k]
                case 'T216.W':
                    for k in range(len(T216W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T216W_data_items[k]
                case 'T217.W':
                    for k in range(len(T217W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T217W_data_items[k]
                case 'T218.R':
                    for k in range(len(T218R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T218R_data_items[k]
                case 'T220.W':
                    for k in range(len(T220W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T220W_data_items[k]
                case 'T222.W':
                    for k in range(len(T222W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T222W_data_items[k]
                case 'T221.R':
                    for k in range(len(T221R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T221R_data_items[k]
                case 'T223.W':
                    for k in range(len(T223W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T223W_data_items[k]
                case 'T224.W':
                    for k in range(len(T224W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T224W_data_items[k]
                case 'T322.W':
                    for k in range(len(T322W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T322W_data_items[k]                              
                case 'T323.W':
                    for k in range(len(T323W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T323W_data_items[k]
                case 'T324.R':
                    for k in range(len(T324R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T324R_data_items[k]
                case 'T325.R':
                    for k in range(len(T325R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T325R_data_items[k]
                case 'T351.R':
                    for k in range(len(T351R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T351R_data_items[k]
                case 'T351.W':
                    for k in range(len(T351W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T351W_data_items[k]
                case 'T352.W':
                    for k in range(len(T352W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T352W_data_items[k]
                case 'T353.R':
                    for k in range(len(T353R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T353R_data_items[k]
                case 'T355.R':
                    for k in range(len(T355R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T355R_data_items[k]
                case 'T355.W':
                    for k in range(len(T355W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T355W_data_items[k]
                case 'T356.W':
                    for k in range(len(T356W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T356W_data_items[k]
                case 'T357.W':
                    for k in range(len(T357W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T357W_data_items[k]
                case 'T365.R':
                    for k in range(len(T365R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T365R_data_items[k]
                case 'T501.R':
                    for k in range(len(T501R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T501R_data_items[k]
                case 'T501.W':
                    for k in range(len(T501W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T501W_data_items[k]
                case 'T505.R':
                    for k in range(len(T505R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T505R_data_items[k]
                case 'T505.W':
                    for k in range(len(T505W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T505W_data_items[k]
                case 'T551.R':
                    for k in range(len(T551R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T551R_data_items[k]
                case 'T551.W':
                    for k in range(len(T551W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T551W_data_items[k]
                case 'T555.R':
                    for k in range(len(T555R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T555R_data_items[k]
                case 'T555.W':
                    for k in range(len(T555W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T555W_data_items[k]
                case 'T556.R':
                    for k in range(len(T556R_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T556R_data_items[k]
                case 'T556.W':
                    for k in range(len(T556W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T556W_data_items[k]
                case 'T557.W':
                    for k in range(len(T557W_data_items)):
                        ws12.cell(row=6+(3*i)+(3*a*TEST_CASE_LENGTH), column=k +
                                7).value = T557W_data_items[k]

    
    test_cases_folder = working_dir + 'TEST_CASES'
    if not os.path.exists(test_cases_folder):
        os.makedirs(test_cases_folder)

    if len(test_case_sequence) > 10:
        new_filename = "RECENT_TESTCASE"
    else:    
        new_filename = '_'.join(test_case_sequence)

    wb1.save(filename = test_cases_folder + '\\' + new_filename.replace('.','') + '.xlsx')

# loop_times repeats test case sequence in the excel file
# max_loop = int (100/TEST_CASE_LENGTH)
max_loop = 1
generate_test_case(max_loop)