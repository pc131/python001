import openpyxl as xl

working_dir = "C:\\Users\\tomasz.skoczylas\\Downloads\\11\\TEST_CASES\\"

bilats_trx_excel = working_dir + "TRX.xlsx"
wb1 = xl.load_workbook(bilats_trx_excel)
ws11 = wb1.worksheets[0]
start_row = 1

while ws11.cell(row = start_row, column = 1).value != "TXXX.W":
    data_items_count = 0
    trx = ws11.cell(row = start_row, column =1 ).value
    # print(trx)
    for col in range(3, 140):
        if ws11.cell(row = start_row, column = col).value != "|": 
            data_items_count += 1
        else:
            break
    print(trx + " Data items count = " + str(data_items_count))
    start_row += 1
