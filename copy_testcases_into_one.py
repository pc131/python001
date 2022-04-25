import openpyxl as xl
import os

working_dir = "C:\\Users\\tomasz.skoczylas\\Documents\\Bi_Laterals\\BiLats_CSD140_CPW070d-B1B3B7\\B7_T365R_RULES\\"
mylist = os.listdir(working_dir)
for filename in mylist:
    if "xlsx" not in filename or "~$TC" in filename:
        mylist.remove(filename)
mylist.remove("Bilaterals 1.4.0.0 master.xlsx")
master_excel = working_dir + "Bilaterals 1.4.0.0 master.xlsx"
wb1 = xl.load_workbook(master_excel)
ws11 = wb1.worksheets[0]
ws12 = wb1.worksheets[1]
total_transactions = 0

for i in range(len(mylist)):
    source_excel = working_dir + "\\" + mylist[i]
    con_rule = "Testing " + mylist[i][8:16]
    wb = xl.load_workbook(source_excel)
    test_case_sequence_sheet = wb.worksheets[0]
    test_case_data_sheet = wb.worksheets[1]
    max_cols = test_case_data_sheet.max_column
    number_of_transactions = 0
    while test_case_sequence_sheet.cell(row = number_of_transactions + 4, column = 3).value != "-":
        trading_party = test_case_sequence_sheet.cell(row = number_of_transactions + 4, column = 3)
        transaction = test_case_sequence_sheet.cell(row = number_of_transactions + 4, column = 5)
        ws11.cell(row = total_transactions + number_of_transactions + 4, column = 3).value = trading_party.value
        ws11.cell(row = total_transactions + number_of_transactions + 4, column = 5).value = transaction.value
        ws11.cell(row = total_transactions + number_of_transactions + 4, column = 7).value = con_rule
        for col in range(7, max_cols+1):
            # reading cell value from source excel file
            value1 = test_case_data_sheet.cell(row = 3*number_of_transactions + 6, column=col)
            # writing the read value to destination excel file
            ws12.cell(row=3*total_transactions + 3*number_of_transactions + 6, column=col).value = value1.value
        number_of_transactions += 1
        total_transactions += 1
wb1.save(filename = working_dir + 'NEWFILE.xlsx')