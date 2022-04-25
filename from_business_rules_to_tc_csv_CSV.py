#### prepare bilaterals most recent business_rules_file (line 22)
#### prepare a valid test script with most recent template_excel_file (line 11) for transaction transaction_name (line 10)
#### test cases for business rules for given transaction will be generated as well as *.csv to import into azure and *.html with urls to execute tests

import openpyxl as xl
import shutil
import os

working_dir = 'C:\\Users\\tomasz.skoczylas\\Downloads\\11\\'
transaction_name = 'T221.R'
# EXCEL WITH CORRECT TRANSACTION
source_test_case = working_dir + 'T353R_T201W_T220W_T221R.xlsx'
transaction_name_simple=transaction_name[1:].replace('.', '')
folder_suffix = '_RULES_TESTCASES'

test_cases_folder = working_dir + 'T' + transaction_name_simple + folder_suffix

if not os.path.exists(test_cases_folder):
    os.makedirs(test_cases_folder)

test_cases_csv = test_cases_folder + '\\IMPORT_T' + transaction_name_simple + '_INTO_AZURE.csv'
urls_for_test_cases = test_cases_folder + '\\T' + transaction_name_simple + '_RULES_TESTCASES_URLS.html'
business_rules_file = working_dir + 'MOSL-Bilaterals-Business-Rules_V0.9.3.xlsx'


wb1 = xl.load_workbook(business_rules_file)
ws12 = wb1.worksheets[0] # Error codes - CHECK IF THIS IS CORRECT SHEET NUMBER!!

#loop through list of transactions in Excel and find column, for current transaction
col_number_trx = 0
#as there are new transactions coming check for maximum column number, where transaction can occur, 52 is for MOSL-Bilaterals-Business-Rules_V0.9.3.xlsx - T551.R
for col_number_trx in range(6, 52):
    if(str(ws12.cell(row=2, column=col_number_trx).value)==transaction_name):
        trx_col_number = col_number_trx

#open CSV file for writing data from Excel
csv_file = open(test_cases_csv, 'w')

#wites a header of CSV file
csv_file.write('ID,Work Item Type,Title,Test Step,Step Action,Step Expected,Area Path,Assigned To,State\n')

#initiate list of elements [[Business Rule 1, Description 1], [Business Rule 2, Description 2], ...]
list_con_xxxx = []
#count number of business rules for current transactiions, basen on column trx_col_number
number_of_business_rules = 0
for row_number1 in range(4, 195): # 195 is the max row, when business rules are defined
    #calculate number of business rules and create list of elements [[Business Rule 1, Description 1], [Business Rule 2, Description 2], ...]
    if(str(ws12.cell(row=row_number1, column=trx_col_number).value)=='X'):
        number_of_business_rules += 1
        #keep adding to the list: ['CON-0005', 'Transaction must be submitted by a Retailer....']
        list_con_xxxx.append([ws12.cell(row=row_number1, column=1).value, ws12.cell(row=row_number1, column=5).value])      



for row_number1 in range(0, number_of_business_rules):

    con_number = list_con_xxxx[row_number1][0] 
    business_rule_description = list_con_xxxx[row_number1][1] 

    #check if decsription can be change according to the dictionary substitute_list[]
    # for element in range(len(substitute_list)):
    #     if substitute_list[element][0] in business_rule_description:
    #         business_rule_description = business_rule_description.replace(substitute_list[element][0], substitute_list[element][1])

    
    csv_file.write(',"Test Case","TC-' + transaction_name_simple + '-' + con_number + '",,,,"Bilaterals UAT Testing","Tomasz Skoczylas <tomasz.skoczylas@cgi.com>","Design"\n')
    csv_file.write(',,,"1","Submit the transaction ' + transaction_name + ' with preconditions against following statement - ' + business_rule_description +'","T209.M rejection notification is issued with ErrorReturnCode ' + con_number + ' and related DataItemReference, to the originator of the transaction",,,\n')

csv_file.close()

#copy template file with correct transaction to files with business rules names to edit
#UNCOMMENT BELOW, when you want to generate *.xlsx files with current transaction to work for specific CON

#for x in range(0, number_of_business_rules):
#   shutil.copy2(template_excel_file, test_cases_folder + '/TC-' + transaction_name_simple + '-' + list_con_xxxx[x][0]  + '.xlsx') 

urls_file = open(urls_for_test_cases, 'w')
urls_file.write('<html>\n<head>\n</head>\n<body>\n')
urls_file.write('<a href=\"https://bilateralhubtestharness-dev-as.azurewebsites.net/api/ExecuteTestCase?code=NbXUVu704AnFSJTiV1JYZ6gq7YCWBqJDxh//C0x/NBAUMnLGWN5uGg==&filename=TEST_SUITE_NAME/T207RW_MOSLTEST_MOSLTEST2.xlsx&format=JSON\" target=\"_blank\">T207RW_MOSLTEST_MOSLTEST2</a><br><br>\n')
for x in range(0, number_of_business_rules):
    print(list_con_xxxx[x][0])
    #GENERATE HTML URLS
    urls_file.write('<a href=\"https://bilateralhubtestharness-dev-as.azurewebsites.net/api/ExecuteTestCase?code=NbXUVu704AnFSJTiV1JYZ6gq7YCWBqJDxh//C0x/NBAUMnLGWN5uGg==&filename=TEST_SUITE_NAME/T' + transaction_name_simple + folder_suffix + '/TC-' + transaction_name_simple + '-' + list_con_xxxx[x][0]  + '.xlsx&format=JSON\" target=\"_blank\">TC-' + transaction_name_simple + '-' + list_con_xxxx[x][0]  + '</a><br><br>\n')
    dest = test_cases_folder +'\\TC-' + transaction_name_simple + '-' + list_con_xxxx[x][0] + '.xlsx'
    # COPY OVER THE XLSX FILES, NEXT WORK ON EVERY TEST CASE (OPEN AND EDIT) TO TRIGGER CORRESPONDING CON
    shutil.copy(source_test_case, dest)
urls_file.write('</body>\n<html>') 
urls_file.close()
