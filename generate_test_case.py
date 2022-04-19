import openpyxl as xl
import random
import os
import sys

# opening the source excel file with first two worksheets
working_dir = "C:\\Users\\tomasz.skoczylas\\Downloads\\11\\"
filename1 = working_dir + "Bilaterals 1.3.3.1 master.xlsx"
wb1 = xl.load_workbook(filename1)
ws11 = wb1.worksheets[0]
ws12 = wb1.worksheets[1]

c1r_transactions = ['T201.W', 'T202.W', 'T203.W', 'T204.R', 'T205.W', 'T206.W', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R',
                    'T211.W', 'T212.W', 'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T321.R', 'T322.W', 'T323.W', 'T324.R', 'T325.R']

C1R_T201W_allowed = ['T203.W', 'T205.W', 'T322.W', 'T323.W']
C1R_T202W_allowed = ['T210.R']
C1R_T203W_allowed = ['T204.R']
C1R_T204R_allowed = ['T203.W', 'T205.W', 'T322.W', 'T323.W']
C1R_T205W_allowed = ['T206.W', 'T212.W', 'T322.W', 'T323.W']
C1R_T206W_allowed = ['T203.W', 'T205.W']
C1R_T210R_allowed = ['T201.W', 'T202.W']
C1R_T212W_allowed = ['T203.W', 'T323.W']
C1R_T321R_allowed = ['T201.W', 'T202.W']
C1R_T322W_allowed = ['T208.R', 'T210.R']
C1R_T323W_allowed = ['T324.R', 'T325.R']
C1R_T324R_allowed = ['T203.W', 'T205.W', 'T322.W']
C1R_T325R_allowed = ['T203.W', 'T323.W']

c1w_transactions = ['T201.W', 'T202.W', 'T205.W', 'T206.W', 'T207.R', 'T207.W', 'T208.R', 'T210.R', 'T211.R', 'T211.W', 'T212.W',
                    'T213.W', 'T214.W', 'T215.R', 'T215.W', 'T216.R', 'T216.W', 'T217.W', 'T218.R', 'T321.W', 'T322.W', 'T323.W', 'T324.R', 'T325.R']

C1W_T201W_allowed = ['T217.W', 'T205.W', 'T322.W', 'T323.W']
C1W_T202W_allowed = ['T210.R']
C1W_T217W_allowed = ['T218.R']
C1W_T218R_allowed = ['T217.W', 'T205.W', 'T322.W', 'T323.W']
C1W_T205W_allowed = ['T206.W', 'T212.W', 'T322.W', 'T323.W']
C1W_T206W_allowed = ['T217.W', 'T205.W']
C1W_T210R_allowed = ['T201.W', 'T202.W']
C1W_T212W_allowed = ['T217.W', 'T323.W']
C1W_T321W_allowed = ['T201.W']
C1W_T322W_allowed = ['T208.R', 'T210.R']
C1W_T323W_allowed = ['T324.R', 'T325.R']
C1W_T324R_allowed = ['T217.W', 'T205.W', 'T322.W']
C1W_T325R_allowed = ['T217.W', 'T323.W']

D8226 = ['NOCONTACT', 'UNCOOPCUST', 'INACCONTACT', 'MOREDETAILS']
D8228 = ['WHOL', 'NONWHOL']
D8230 = ['INACCURATE', 'DUPLICATE', 'WRONGPRO', 'POLICY', 'HOUSEHOLD', 'NOTWHOL']
D8231 = ['DISPREJECT', 'DISPCMOS']

T321R_data_items = ['3200662662W19', 'MEASURED', 'RET1234', '', '1', '[today]', '1',
                    # customer and retailer data
                    'Customer1', '4431244526', '12', 'Customer2', '4432122345', '13', 'email@email.co.uk', '1', 'EMAIL', 'AM',
                    'More info', 'NA', 'Details', 'Retailer1', '441234545', '12', 'Retailer2', '443451213', '13', 'ret_email@ret.co.uk',
                    # first meter details
                    'ELSTER', '10W11171016329', '0', '120', '[today-1]', '0', 'METER', '1', 'ELSTER3', '1', '10W12345123456', '1', '5', '1', '5',
                    '1', '82650', '1', '10000', '1', 'O', '1', 'Under the tree', '1', '82650', '1', '10000', '1', 'I', '1', 'Somewhere', 'more info', '',
                    # second meter details empty
                    '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                    # missing meters
                    '', '', '', '',
                    # unmeasrued data empty
                    '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''
                    ]
T201W_data_items = ['[orid]', 'WSL-T987GHD52J']
T202W_data_items = ['[orid]', 'WSL-T987GHD52J', random.choice(D8230), 'More info']
T203W_data_items = ['[orid]', random.choice(D8226), 'More info']
T204R_data_items = ['[orid]', 'More info']
T205W_data_items = ['[orid]', '[today+5]', '', 'More info']
T206W_data_items = ['[orid]', random.choice(D8228), 'More info']
T208R_data_items = ['[orid]', 'CLOSING']
T210R_data_items = ['[orid]', random.choice(D8231), 'More info 1234']
T212W_data_items = ['[orid]', 'Preparing plan']
T323W_data_items = ['[orid]', 'ABLE', '[now+3]',
                    '0', '1', '[today+3]', 'Proposed a plan']
T324R_data_items = ['[orid]', 'AGREED a plan']
T325R_data_items = ['[orid]', 'DISPUTING a plan']
T322W_data_items = ['[orid]', '1', 'ELSTER', '10W11171016329', '0', '1', 'ELSTER4', '1', '10W11171000000', '1', '12', '1', '5',
                    '1', '82652', '1', '10001', '1', 'I', '1', 'SOMEWHERE', '1', '82652', '1', '10003', '1', 'O', '1', 'ON THE ROOF', 'MORE INFO 322',
                    '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                    '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']


def generate_test_case(number_of_steps):
    last_transaction = False
    test_case_sequence = ['T321.R']
    # print('\nTest case sequence: ', test_case_sequence)
    last_transaction = test_case_sequence[-1]
    for i in range(1, number_of_steps):
        if last_transaction !=True:
            transactions_allowed = globals(
            )['C1R_' + last_transaction.replace('.', '') + '_allowed']
            print('last transaction: '+str(test_case_sequence[-1]))
            print('transactions allowed: ', transactions_allowed)
            print('i= ' + str(i))
            next_tran = random.choice(transactions_allowed)
            print('next transaction chosen: ' + next_tran)
            test_case_sequence.append(next_tran)
            last_transaction = test_case_sequence[-1]
            print('Test case sequence: ',test_case_sequence)
            if next_tran == 'T208.R':
                last_transaction = True

    print('Test case sequence: ', test_case_sequence)
    new_filename = ''
    #now loop as many times, as there are trancastcions in the test case - in case T208.R was selected earlier...
    for i in range(len(test_case_sequence)):
        ws11.cell(row=i+4, column=5).value = test_case_sequence[i]
        new_filename = new_filename + test_case_sequence[i] + '_'
        if test_case_sequence[i][-1] == 'R':
            ws11.cell(row=i+4, column=3).value = 'MOSLTEST-R'
        else:
            ws11.cell(row=i+4, column=3).value = 'MOSLTEST-W'

        match test_case_sequence[i]:
            case 'T321.R':
                for k in range(len(T321R_data_items)):
                    ws12.cell(row=6+(3*i), column=k +
                              7).value = T321R_data_items[k]
            case 'T201.W':
                for k in range(len(T201W_data_items)):
                    ws12.cell(row=6+(3*i), column=k +
                              7).value = T201W_data_items[k]
            case 'T202.W':
                for k in range(len(T202W_data_items)):
                    ws12.cell(row=6+(3*i), column=k +
                              7).value = T202W_data_items[k]
            case 'T203.W':
                for k in range(len(T203W_data_items)):
                    ws12.cell(row=6+(3*i), column=k +
                              7).value = T203W_data_items[k]
            case 'T204.R':
                for k in range(len(T204R_data_items)):
                    ws12.cell(row=6+(3*i), column=k +
                              7).value = T204R_data_items[k]
            case 'T205.W':
                for k in range(len(T205W_data_items)):
                    ws12.cell(row=6+(3*i), column=k +
                              7).value = T205W_data_items[k]
            case 'T206.W':
                for k in range(len(T206W_data_items)):
                    ws12.cell(row=6+(3*i), column=k +
                              7).value = T206W_data_items[k]
            case 'T208.R':
                for k in range(len(T208R_data_items)):
                    ws12.cell(row=6+(3*i), column=k +
                              7).value = T208R_data_items[k]
            case 'T210.R':
                for k in range(len(T210R_data_items)):
                    ws12.cell(row=6+(3*i), column=k +
                              7).value = T210R_data_items[k]
            case 'T212.W':
                for k in range(len(T212W_data_items)):
                    ws12.cell(row=6+(3*i), column=k +
                              7).value = T212W_data_items[k]
            case 'T322.W':
                for k in range(len(T322W_data_items)):
                    ws12.cell(row=6+(3*i), column=k +
                              7).value = T322W_data_items[k]                              
            case 'T323.W':
                for k in range(len(T323W_data_items)):
                    ws12.cell(row=6+(3*i), column=k +
                              7).value = T323W_data_items[k]
            case 'T324.R':
                for k in range(len(T324R_data_items)):
                    ws12.cell(row=6+(3*i), column=k +
                              7).value = T324R_data_items[k]
            case 'T325.R':
                for k in range(len(T325R_data_items)):
                    ws12.cell(row=6+(3*i), column=k +
                              7).value = T325R_data_items[k]
    #remove last char from a string
    new_filename = new_filename[:-1]
    wb1.save(filename = working_dir + new_filename.replace('.','') + '.xlsx')


generate_test_case(15)
