import openpyxl as xl
import os
# list files in current directory
mylist = os.listdir("C:\\Users\\tomasz.skoczylas\\Downloads\\11\\TEST_CASES\\")
# remove the template file from the list, so it is not used as source file
mylist.remove("Bilaterals 1.8.0.1 master.xlsx")

for f in mylist:

    filename1 = "C:\\Users\\tomasz.skoczylas\\Downloads\\11\\TEST_CASES\\" + f
    # load the excel file - workbook
    wb1 = xl.load_workbook(filename1)
    # work on worksheet 2 - Test Data
    ws10 = wb1.worksheets[0]
    ws11 = wb1.worksheets[1]
    #look for specific transaction in file and change it to thter transaction, change also values data
    rown_nmbr = 0
    for i in range (4, 14):
        print('Value of cell = ' + ws10.cell(row=i, column=5).value)
        if ws10.cell(row=i, column=5).value == 'T225.R':
            rown_nmbr = i - 3
    print('Row number = ' + str(rown_nmbr))
    #update 6th row, AW column in excel, worksheet 2 (Test Data)
    #ws11.cell(row=6, column=22).value = ""
    #ws11.cell(row=6, column=23).value = "METERED"
    #ws11.cell(row=6, column=24).value = ""
    j = 3 + rown_nmbr*3
    row1 = rown_nmbr + 3
    ws10.cell(row=row1, column=3).value = 'MOSLTEST-W'
    ws10.cell(row=row1, column=5).value = 'T225.W'
    ws11.cell(row=j, column=8).value = 'INCORRECTRTL'
    ws11.cell(row=j, column=9).value = 'ADDITIONAL T225W'

    # # save updated file
    wb1.save(filename1)