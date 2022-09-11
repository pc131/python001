from ast import match_case
import openpyxl as xl
import shutil
import os

working_dir = 'C:\\Users\\tomasz.skoczylas\\Downloads\\11\\'
filename = 'TESTING CANDIDATES - Sprint 50_04072022.xlsx'
source_xls = working_dir + filename
sprint_xls_csv = working_dir + filename.replace('xlsx','csv')
wb1 = xl.load_workbook(source_xls)
ws1 = wb1.worksheets[0]
csv_file = open(sprint_xls_csv, 'w')

csv_file.write('ID,Work Item Type,Title,Test Step,Step Action,Step Expected,Area Path,Assigned To,State\n')
for row_number in range(1, 20):
    id = ws1.cell(row=row_number, column=1).value
    id_type = ws1.cell(row=row_number, column=2).value
    if id_type == 'Bug':
        id_type_1 = 'BUG'
    else:
        id_type_1 = 'PBI'
    id_desc = ws1.cell(row=row_number, column=3).value
    id_desc1 = id_desc.replace('"','')
    specialChars = '/.:'
    for specialChar in specialChars:
        id_desc_2 = id_desc1.replace(specialChar, '_')
    id_desc_3 = id_desc_2.replace(' ', '_')[:35]
    csv_file.write(',"Test Case","' + id_type_1 + '_' + str(id) + '_' + id_desc_3 + '",,,,"Bilaterals UAT Testing","Tomasz Skoczylas <tomasz.skoczylas@cgi.com>","Design"\n')
    csv_file.write(',,,"1","Prepare and execute a test case to ckeck following business conditions: ' + id_desc1 +'","Test case was executed and results are as expected",,,\n')

csv_file.close()