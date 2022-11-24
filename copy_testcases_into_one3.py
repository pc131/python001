import openpyxl as xl
import os

TRX_DATA_TEMS_COUNT = {'T201.W':'2','T202.W':'4','T203.W':'3','T204.R':'2','T205.W':'4','T206.W':'3','T207.R':'2','T207.W':'2','T208.R':'2','T210.R':'3','T211.R':'3','T211.W':'3','T212.W':'2','T213.W':'5','T214.W':'3','T215.R':'5','T215.W':'5','T216.R':'2','T216.W':'2','T217.W':'4','T218.R':'25','T220.W':'2','T221.R':'2','T222.W':'9','T223.W':'26','T224.W':'3','T225.R':'4','T225.W':'4','T226.W':'3','T227.R':'2','T228.R':'2','T321.R':'133','T321.W':'110','T322.W':'91','T323.W':'7','T324.R':'2','T325.R':'2','T331.W':'24','T332.W':'29','T335.R':'76','T336.W':'64','T341.R':'102','T341.W':'79','T342.W':'11','T351.R':'53','T351.W':'30','T352.W':'29','T353.R':'42','T355.R':'48','T355.W':'25','T356.W':'12','T357.W':'28','T365.R':'51','T501.R':'35','T501.W':'12','T505.R':'31','T505.W':'8','T551.R':'30','T551.W':'7','T555.R':'39','T555.W':'16','T556.R':'50','T556.W':'28','T557.W':'5','T561.R':'44','T561.W':'21','T562.R':'57','T562.W':'34','T563.W':'5'}

# copies transactions from multiple excel files into 1 excel file with these transactions
working_dir = "C:\\Users\\tomasz.skoczylas\\Downloads\\11\\TEST_CASES\\"
mylist = os.listdir(working_dir)
# for filename in mylist:
#     if "CON-" not in filename or "z" in filename:
#         mylist.remove(filename)
mylist.remove("Bilaterals 1.8.0.1 master.xlsx")
master_excel = working_dir + "Bilaterals 1.8.0.1 master.xlsx"
wb1 = xl.load_workbook(master_excel)
ws11 = wb1.worksheets[0]
ws12 = wb1.worksheets[1]
total_transactions = 0

#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected


#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):

            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

for a in mylist:
    print(a)

for i in range(len(mylist)):
    source_excel = working_dir + "\\" + mylist[i]
    con_rule = "Testing " + mylist[i][8:16]
    wb = xl.load_workbook(source_excel)

    test_case_sequence_sheet = wb.worksheets[0]
    test_case_data_sheet = wb.worksheets[1]

    number_of_transactions = 0
    while test_case_sequence_sheet.cell(row = number_of_transactions + 4, column = 3).value != "-":
        print("\nWorking with file number: " + str(i+1) + " " + source_excel)
        number_of_transactions += 1
        trading_party = test_case_sequence_sheet.cell(row = number_of_transactions + 3, column = 3)
        transaction = test_case_sequence_sheet.cell(row = number_of_transactions + 3, column = 5)
        ws11.cell(row = total_transactions + number_of_transactions + 3, column = 3).value = trading_party.value
        ws11.cell(row = total_transactions + number_of_transactions + 3, column = 5).value = transaction.value

        # selectedRange = copyRange(7,3*(number_of_transactions-1) + 6,139,3*(number_of_transactions-1) + 6,test_case_data_sheet) #Change the 4 number values
        # pastingRange = pasteRange(7,3*(total_transactions-1) + 3*(number_of_transactions) + 6,139,3*(total_transactions-1) + 3*(number_of_transactions) + 6,ws12,selectedRange) #Change the 4 number values
 
        value1 = test_case_data_sheet.cell(row = 3*(number_of_transactions-1) + 6, column=22)
        # writing the read value to destination excel file
        print('Reading value ' + str(value1.value) + ' from cell ' + str(3*(number_of_transactions-1) + 6) + ' ' + str(22))
        print('Writing value ' + str(value1.value) + ' for cell ' + str(3*(total_transactions) + 3*(number_of_transactions-1) + 6) + ' ' + str(22))
        ws12.cell(row=3*(total_transactions) + 3*(number_of_transactions-1) + 6, column=22).value = value1.value

    total_transactions += number_of_transactions

    print("Currently total transactions in all files: " + str(total_transactions))
    print("----------------------")
    print("----------------------")
wb1.save(filename = working_dir + 'NEWFILE.xlsx')