import openpyxl as xl
import os
# list files in current directory
mylist = os.listdir("C:\\Users\\tomasz.skoczylas\\Downloads\\11\\")
# remove the template file from the list, so it is not used as source file
mylist.remove('Bilaterals 1.1 master.xlsx')

for f in mylist:

    # opening the source excel file with first two worksheets
    filename1 = "C:\\Users\\tomasz.skoczylas\\Downloads\\11\\" + f
    wb1 = xl.load_workbook(filename1)
    ws11 = wb1.worksheets[0]
    ws12 = wb1.worksheets[1]

    # opening the destination excel file
    filename2 = "C:\\Users\\tomasz.skoczylas\\Downloads\\11\\Bilaterals 1.1 master.xlsx"
    wb2 = xl.load_workbook(filename2)
    ws21 = wb2.worksheets[0]
    ws22 = wb2.worksheets[1]

    # calculate total number of rows and columns in source excel file worksheet1
    mrws11 = ws11.max_row
    mcws12 = ws11.max_column
    #print (str(mrws11)+str(mcws12))
    # calculate total number of rows and columns in source excel file worksheet2
    mrws21 = ws21.max_row
    mcws22 = ws22.max_column

    # copying the cell values from source excel file worksheet1 to destination excel file worksheet1

    for i in range(1, mrws11 + 1):
        for j in range(1, mcws12 + 1):
            # reading cell value from source excel file
            c = ws11.cell(row=i, column=j)
            # writing the read value to destination excel file
            ws21.cell(row=i, column=j).value = c.value
    # add info about version
    ws21.cell(
        row=4, column=7).value = 'created with 1.1 excel template'

    # copying the cell values from source excel file worksheet2 to destination excel file worksheet2
    for k in range(1, mrws21 + 1):
        for l in range(1, mcws22 + 1):
            # reading cell value from source excel file
            d = ws12.cell(row=k, column=l)
            # writing the read value to destination excel file
            ws22.cell(row=k, column=l).value = d.value
    
    # saving the destination excel file and replace version in filename (092 to 102)
    #wb2.save(str(filename1.replace("092", "102")))
    wb2.save(str(filename1.replace(".xlsx", "_new.xlsx")))