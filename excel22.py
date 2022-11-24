import os
import pandas as pd

from openpyxl import load_workbook

working_dir = "C:\\Users\\tomasz.skoczylas\\Downloads\\11\\TEST_CASES\\"

wb = load_workbook(filename = working_dir + 'TC-332W-CON-0004.xlsx',
                   data_only=True)

ws = wb['Test Case sequence']

def range_to_df(ws, remove_nan=True):
    # Read the cell values into a list of lists
    data_rows = []
    for row in ws:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)
    df = pd.DataFrame(data_rows[0:])
    df.columns = ['ID', 'TP','-', 'TRX']
    if remove_nan:
        df.dropna(axis=1, how='all', inplace=True)
    print(df)

ethnicity = range_to_df(ws['B4':'E103'])