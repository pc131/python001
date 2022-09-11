import openpyxl as xl
import os
# list files in current directory
mylist = os.listdir("C:\\Users\\tomasz.skoczylas\\Downloads\\11\\TEST_CASES\\")
# remove the template file from the list, so it is not used as source file
mylist.remove("Bilaterals 1.6.0.1 master.xlsx")

for f in mylist:

    filename1 = "C:\\Users\\tomasz.skoczylas\\Downloads\\11\\TEST_CASES\\" + f
    # load the excel file - workbook
    wb1 = xl.load_workbook(filename1)
    # work on worksheet 2 - Test Data
    ws11 = wb1.worksheets[1]
    # update 6th row, AW column in excel, worksheet 2 (Test Data)
    #ws11.cell(row=6, column=22).value = ""
    #ws11.cell(row=6, column=23).value = "METERED"
    #ws11.cell(row=6, column=24).value = ""
    ws11.cell(row=6, column=61).value = ""
    # save updated file
    wb1.save(filename1)